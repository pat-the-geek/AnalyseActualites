"""
WUDD.ai Viewer — Flask backend
Sert l'API de navigation de fichiers et le frontend React compilé.
"""

import json
import os
import re
import subprocess
import threading
import datetime
from pathlib import Path
from flask import Flask, jsonify, send_file, request, abort, send_from_directory, Response, stream_with_context

app = Flask(__name__)

# Suivi du process RSS keyword (un seul à la fois)
_rss_job: dict = {
    "process": None,
    "lock": threading.Lock(),
    "last_run": None,        # ISO 8601 UTC — horodatage de la dernière fin d'exécution
    "last_returncode": None, # Code retour de la dernière exécution
}

# Charge les variables d'environnement depuis .env (si disponible)
try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv(Path(__file__).resolve().parent.parent / ".env", override=False)
except ImportError:
    pass

# La racine du projet est le dossier parent de viewer/
# resolve() AVANT parent.parent : __file__ peut être un chemin relatif
# ('app.py') quand Flask est lancé via `python3 app.py` depuis viewer/,
# auquel que (Path('app.py').parent.parent).resolve() → cwd au lieu de la racine.
PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Ajouter la racine au sys.path pour que `from utils.X import Y` fonctionne
# quel que soit le répertoire courant au démarrage (cron, Docker, CLI…)
import sys as _sys
if str(PROJECT_ROOT) not in _sys.path:
    _sys.path.insert(0, str(PROJECT_ROOT))

from utils.api_client import CLAUDE_API_VERSION


# ── Utilitaires ───────────────────────────────────────────────────────────────

def safe_path(relative: str) -> Path:
    """Résout et valide un chemin pour qu'il reste dans PROJECT_ROOT."""
    resolved = (PROJECT_ROOT / relative).resolve()
    if not str(resolved).startswith(str(PROJECT_ROOT) + "/") and resolved != PROJECT_ROOT:
        abort(403, "Accès refusé")
    if not resolved.exists():
        abort(404, "Fichier non trouvé")
    return resolved


def collect_files() -> list:
    files = []

    def scan(directory: Path, file_type: str, flux_override: str | None = None):
        if not directory.exists():
            return
        exts = ["*.json"] if file_type == "json" else ["*.md", "*.markdown"]
        for ext in exts:
            # Matérialiser le générateur rglob() pour intercepter les OSError
            # survenant pendant le parcours (volumes Docker, race avec cron)
            try:
                paths = list(directory.rglob(ext))
            except OSError:
                paths = []
            for f in paths:
                parts = f.relative_to(PROJECT_ROOT).parts
                if any(p in ("cache", ".git") for p in parts):
                    continue
                try:
                    stat = f.stat()
                    rel = f.relative_to(PROJECT_ROOT)
                    flux = flux_override or (f.parent.name if f.parent != directory else "")
                    files.append({
                        "name": f.name,
                        "path": str(rel).replace("\\", "/"),
                        "type": file_type,
                        "flux": flux,
                        "size": stat.st_size,
                        "modified": stat.st_mtime,
                    })
                except OSError:
                    continue

    # Scan large de data/ : couvre articles/, articles-from-rss/, et toute
    # autre structure que l'utilisateur pourrait avoir sous data/
    scan(PROJECT_ROOT / "data", "json")
    scan(PROJECT_ROOT / "rapports", "markdown")
    # Fichiers d'exemple (visibles tant que data/ et rapports/ sont vides)
    scan(PROJECT_ROOT / "samples", "json",     "Samples")
    scan(PROJECT_ROOT / "samples", "markdown", "Samples")
    return sorted(files, key=lambda x: x["modified"], reverse=True)


def parse_cron_field(s: str, lo: int, hi: int) -> list:
    if s == "*":
        return list(range(lo, hi + 1))
    if s.startswith("*/"):
        step = int(s[2:])
        return list(range(lo, hi + 1, step))
    if "," in s:
        return [int(x) for x in s.split(",")]
    if "-" in s:
        # Gère a-b/step (ex: 6-22/2) et a-b
        if "/" in s:
            rng, step_str = s.split("/", 1)
            a, b = rng.split("-", 1)
            return list(range(int(a), int(b) + 1, int(step_str)))
        a, b = s.split("-", 1)
        return list(range(int(a), int(b) + 1))
    return [int(s)]


def next_cron_occurrence(cron: str, after: datetime.datetime | None = None) -> datetime.datetime | None:
    """Calcule la prochaine occurrence d'une expression cron à 5 champs."""
    if after is None:
        after = datetime.datetime.now()
    parts = cron.strip().split()
    if len(parts) != 5:
        return None
    try:
        minutes = parse_cron_field(parts[0], 0, 59)
        hours   = parse_cron_field(parts[1], 0, 23)
        doms    = parse_cron_field(parts[2], 1, 31)
        months  = parse_cron_field(parts[3], 1, 12)
        # cron DOW: 0=dim…6=sam ; Python isoweekday: 1=lun…7=dim
        if parts[4] == "*":
            dows = set(range(1, 8))
        else:
            raw = parse_cron_field(parts[4], 0, 7)
            # 0 et 7 = dimanche (isoweekday=7)
            dows = set(7 if d in (0, 7) else d for d in raw)

        current = after.replace(second=0, microsecond=0) + datetime.timedelta(minutes=1)
        end = after + datetime.timedelta(days=35)

        while current <= end:
            if (current.month in months
                    and current.day in doms
                    and current.isoweekday() in dows
                    and current.hour in hours
                    and current.minute in minutes):
                return current
            # Optimisation : avancer directement à la bonne minute
            valid_mins = sorted(m for m in minutes if m > current.minute)
            if not valid_mins:
                current = current.replace(minute=0) + datetime.timedelta(hours=1)
            else:
                current = current.replace(minute=valid_mins[0])
        return None
    except (ValueError, IndexError):
        return None


def cron_label(cron: str) -> str:
    """Description humaine en français d'une expression cron courante."""
    p = cron.strip().split()
    if len(p) != 5:
        return cron
    minute, hour, dom, month, dow = p
    jours = ["Dimanche", "Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]

    # Toutes les N minutes : */N * * * *
    if minute.startswith("*/"):
        return f"Toutes les {minute[2:]} min"

    # Toutes les Xh entre heures : 0 6-22/2 * * *
    if minute == "0" and "/" in hour and "-" in hour and dom == "*":
        try:
            range_part, step = hour.split("/")
            start, end = range_part.split("-")
            return f"Toutes les {step}h de {start}h à {end}h"
        except ValueError:
            pass

    # Heure H:M — construire le libellé d'heure
    if minute.isdigit() and hour.isdigit():
        t = f"{int(hour):02d}:{int(minute):02d}"

        # Fin de mois (dom 28-31) : M H 28-31 * *
        if dom == "28-31" and month == "*" and dow == "*":
            return f"Fin de mois à {t}"

        # Jour de semaine spécifique, tous les mois : M H * * D
        if dom == "*" and month == "*" and dow.isdigit():
            try:
                return f"{jours[int(dow) % 7]} à {t}"
            except (ValueError, IndexError):
                pass

        # Quotidien : M H * * *
        if dom == "*" and month == "*" and dow == "*":
            return f"Quotidien à {t}"

    return cron


def latest_mtime(directory: Path) -> datetime.datetime | None:
    """Retourne la date de modification du fichier JSON le plus récent dans directory."""
    if not directory.exists():
        return None
    candidates = [
        f for f in directory.rglob("*.json")
        if "cache" not in f.relative_to(directory).parts
    ]
    if not candidates:
        return None
    return datetime.datetime.fromtimestamp(max(f.stat().st_mtime for f in candidates))


# ── Routes API ────────────────────────────────────────────────────────────────

@app.route("/api/files")
def api_files():
    import time
    # Double scan pour compenser les listings incomplets de virtiofs
    # (Docker Desktop / macOS) : rglob() peut retourner un résultat partiel
    # si le cache kernel est rafraîchi entre les deux appels.
    files1 = collect_files()
    time.sleep(0.20)
    files2 = collect_files()
    # Union des deux passes — chaque chemin vu dans l'une ou l'autre est retenu ;
    # la seconde passe écrase les métadonnées de la première si les deux la voient.
    by_path = {f["path"]: f for f in files1}
    by_path.update({f["path"]: f for f in files2})
    files = sorted(by_path.values(), key=lambda x: x["modified"], reverse=True)
    return jsonify(files)


@app.route("/api/content")
def api_content():
    path = request.args.get("path", "")
    if not path:
        abort(400)
    f = safe_path(path)
    return jsonify({"path": path, "content": f.read_text(encoding="utf-8", errors="replace")})


@app.route("/api/stream-content")
def api_stream_content():
    """Diffuse le contenu d'un fichier en streaming pour une meilleure réactivité."""
    path = request.args.get("path", "")
    if not path:
        abort(400)
    f = safe_path(path)
    file_size = f.stat().st_size

    def generate():
        with open(f, "rb") as fh:
            while True:
                chunk = fh.read(16384)  # 16 Ko par chunk
                if not chunk:
                    break
                yield chunk

    return Response(
        stream_with_context(generate()),
        mimetype="text/plain; charset=utf-8",
        headers={
            "X-File-Size": str(file_size),
            "Cache-Control": "no-cache",
        },
    )


@app.route("/api/content", methods=["POST"])
def api_save_content():
    data = request.get_json(force=True)
    if not data or "path" not in data or "content" not in data:
        abort(400, "Champs 'path' et 'content' requis")
    rel = data["path"]
    # Restriction : uniquement data/ et config/ sont modifiables
    if not (rel.startswith("data/") or rel.startswith("config/")):
        abort(403, "Modification non autorisée hors de data/ et config/")
    target = (PROJECT_ROOT / rel).resolve()
    if not str(target).startswith(str(PROJECT_ROOT) + "/"):
        abort(403, "Accès refusé")
    if not target.exists():
        abort(404, "Fichier non trouvé")
    # Validation JSON si extension .json
    content = data["content"]
    if target.suffix == ".json":
        try:
            json.loads(content)
        except json.JSONDecodeError as e:
            abort(400, f"JSON invalide : {e}")
    try:
        target.write_text(content, encoding="utf-8")
    except OSError as e:
        abort(500, f"Erreur d'écriture : {e}")
    return jsonify({"ok": True})


@app.route("/api/search")
def api_search():
    """Recherche textuelle avec filtres optionnels.

    Paramètres :
      q          : texte à chercher (min 2 chars)
      type       : "json" ou "markdown" (filtre type de fichier)
      sentiment  : "positif", "neutre", "négatif" (filtre sur articles JSON)
      source     : nom de source (filtre partiel, insensible à la casse)
      date_from  : YYYY-MM-DD (articles publiés à partir de cette date)
      date_to    : YYYY-MM-DD (articles publiés jusqu'à cette date)
    """
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])

    filter_type     = request.args.get("type", "").strip().lower()
    filter_sentiment = request.args.get("sentiment", "").strip().lower()
    filter_source   = request.args.get("source", "").strip().lower()
    filter_from     = request.args.get("date_from", "").strip()
    filter_to       = request.args.get("date_to", "").strip()

    has_article_filters = bool(filter_sentiment or filter_source or filter_from or filter_to)

    pattern = re.compile(re.escape(q), re.IGNORECASE)
    results = []

    for info in collect_files():
        # Filtre par type de fichier
        if filter_type and info["type"] != filter_type:
            continue

        f = PROJECT_ROOT / info["path"]

        # Pour les filtres article, on parse le JSON et applique les filtres
        if has_article_filters and info["type"] == "json":
            try:
                articles = json.loads(f.read_text(encoding="utf-8", errors="replace"))
                if not isinstance(articles, list):
                    continue
            except (json.JSONDecodeError, OSError):
                continue

            # Filtrer les articles selon les critères
            filtered = []
            for art in articles:
                if filter_sentiment and art.get("sentiment", "").lower() != filter_sentiment:
                    continue
                if filter_source:
                    src = art.get("Sources", "").lower()
                    if filter_source not in src:
                        continue
                date_str = art.get("Date de publication", "")[:10]
                if filter_from and date_str and date_str < filter_from:
                    continue
                if filter_to and date_str and date_str > filter_to:
                    continue
                # Vérifier si la requête textuelle matche
                resume = art.get("Résumé", "") or ""
                if pattern.search(resume) or pattern.search(art.get("URL", "") or "") or pattern.search(art.get("Sources", "") or ""):
                    filtered.append(art)

            if not filtered:
                continue

            matches = [
                {"line": 0, "text": f"{art.get('Sources','')} · {art.get('Date de publication','')[:10]} — {(art.get('Résumé','') or '')[:150]}"}
                for art in filtered[:5]
            ]
            results.append({**info, "matches": matches, "article_count": len(filtered)})

        else:
            # Recherche ligne par ligne (fichiers Markdown ou JSON sans filtres article)
            try:
                lines = f.read_text(encoding="utf-8", errors="replace").splitlines()
                matches = [
                    {"line": i + 1, "text": line.strip()[:200]}
                    for i, line in enumerate(lines)
                    if pattern.search(line)
                ]
                if matches:
                    results.append({**info, "matches": matches[:5]})
            except OSError:
                continue

    return jsonify(results)


@app.route("/api/download")
def api_download():
    path = request.args.get("path", "")
    if not path:
        abort(400)
    f = safe_path(path)
    return send_file(f, as_attachment=True, download_name=f.name)


@app.route("/api/scheduler")
def api_scheduler():
    now = datetime.datetime.now()
    tasks = []

    # Lire l'état du flux_watcher pour afficher le flux en cours de traitement
    flux_watcher_state_file = PROJECT_ROOT / "data" / "flux_watcher_state.json"
    flux_watcher_detail = None
    flux_watcher_last_run = None
    if flux_watcher_state_file.exists():
        try:
            _fw_state = json.loads(flux_watcher_state_file.read_text(encoding="utf-8"))
            last_idx   = _fw_state.get("last_feed_idx", 0)
            feed_count = _fw_state.get("feed_count", 0)
            last_title = _fw_state.get("last_feed_title", "")
            added      = _fw_state.get("articles_added", 0)
            next_idx   = (last_idx + 1) % feed_count if feed_count > 0 else 0
            flux_watcher_detail = (
                f"Dernier flux : [{last_idx + 1}/{feed_count}] {last_title}"
                f" (+{added} articles) — Prochain : #{next_idx + 1}"
            )
            _fw_last = _fw_state.get("last_run")
            if _fw_last:
                try:
                    flux_watcher_last_run = datetime.datetime.fromisoformat(_fw_last).replace(tzinfo=None)
                except Exception:
                    pass
        except Exception:
            pass

    # Tâches cron fixes (issues de archives/crontab)
    # Catégories : "Surveillance en continu" | "Enrichissement nocturne" | "Rapports & digests" | "Pipeline mensuel"
    fixed = [
        # ── Surveillance en continu ──────────────────────────────────────────────
        {
            "name": "Veille RSS temps-réel (round-robin)",
            "script": "flux_watcher.py → entity_timeline.py → cross_flux_analysis.py → enrich_reading_time.py",
            "cron": "*/5 * * * *",
            "category": "Surveillance en continu",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_flux_watcher.log",
            "extra_last_run": flux_watcher_last_run,
            "detail": flux_watcher_detail,
        },
        {
            "name": "Surveillance sources web (sitemap)",
            "script": "web_watcher.py",
            "cron": "0 */2 * * *",
            "category": "Surveillance en continu",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_web_watcher.log",
        },
        {
            "name": "Extraction mots-clés RSS",
            "script": "get-keyword-from-rss.py",
            "cron": "0 6-22/2 * * *",
            "category": "Surveillance en continu",
            "data_dir": PROJECT_ROOT / "data" / "articles-from-rss",
        },
        {
            "name": "Vérification santé cron",
            "script": "check_cron_health.py",
            "cron": "*/10 * * * *",
            "category": "Surveillance en continu",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_health.log",
        },
        # ── Enrichissement nocturne ──────────────────────────────────────────────
        {
            "name": "Backup des données",
            "script": "backup_data.py",
            "cron": "0 1 * * *",
            "category": "Enrichissement nocturne",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_backup.log",
        },
        {
            "name": "Enrichissement NER (entités)",
            "script": "enrich_entities.py",
            "cron": "0 2 * * *",
            "category": "Enrichissement nocturne",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_enrich_entities.log",
        },
        {
            "name": "Enrichissement images",
            "script": "enrich_images.py",
            "cron": "30 2 * * *",
            "category": "Enrichissement nocturne",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_enrich_images.log",
        },
        {
            "name": "Enrichissement sentiment",
            "script": "enrich_sentiment.py",
            "cron": "0 3 * * *",
            "category": "Enrichissement nocturne",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_sentiment.log",
        },
        {
            "name": "Réparation résumés en erreur",
            "script": "repair_failed_summaries.py",
            "cron": "0 4 * * 0",
            "category": "Enrichissement nocturne",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_repair.log",
        },
        # ── Rapports & digests ───────────────────────────────────────────────────
        {
            "name": "Collecte multi-flux",
            "script": "scheduler_articles.py",
            "cron": "0 6 * * 1",
            "category": "Rapports & digests",
            "data_dir": PROJECT_ROOT / "data" / "articles",
        },
        {
            "name": "Briefing exécutif hebdomadaire",
            "script": "generate_briefing.py --period weekly",
            "cron": "30 6 * * 1",
            "category": "Rapports & digests",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_briefing.log",
        },
        {
            "name": "Détection tendances & alertes",
            "script": "trend_detector.py",
            "cron": "0 7 * * *",
            "category": "Rapports & digests",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_trends.log",
        },
        {
            "name": "Morning Digest quotidien",
            "script": "generate_morning_digest.py --ai",
            "cron": "30 7 * * *",
            "category": "Rapports & digests",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_morning_digest.log",
        },
        {
            "name": "Notes de lecture quotidiennes",
            "script": "generate_reading_notes.py",
            "cron": "0 8 * * *",
            "category": "Rapports & digests",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_reading_notes.log",
        },
        {
            "name": "Rapport Top 10 entités 48h",
            "script": "generate_48h_report.py",
            "cron": "0 23 * * *",
            "category": "Rapports & digests",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_48h_report.log",
        },
        # ── Pipeline mensuel ─────────────────────────────────────────────────────
        {
            "name": "Radar thématique",
            "script": "radar_wudd.py",
            "cron": "0 5 28-31 * *",
            "category": "Pipeline mensuel",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_radar.log",
        },
        {
            "name": "Conversion articles RSS → Markdown",
            "script": "articles_rss_to_markdown.py",
            "cron": "30 5 28-31 * *",
            "category": "Pipeline mensuel",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_rss_markdown.log",
        },
        {
            "name": "Rapports mensuels par mot-clé",
            "script": "generate_keyword_reports.py",
            "cron": "0 6 28-31 * *",
            "category": "Pipeline mensuel",
            "data_dir": None,
            "log_file": PROJECT_ROOT / "rapports" / "cron_keyword_reports.log",
        },
    ]
    for t in fixed:
        if t.get("extra_last_run"):
            last_run = t["extra_last_run"]
        elif t.get("data_dir"):
            last_run = latest_mtime(t["data_dir"])
        elif t.get("log_file") and t["log_file"].exists():
            last_run = datetime.datetime.fromtimestamp(t["log_file"].stat().st_mtime)
        else:
            last_run = None
        next_run = next_cron_occurrence(t["cron"], now)
        tasks.append({
            "name": t["name"],
            "script": t["script"],
            "cron": t["cron"],
            "label": cron_label(t["cron"]),
            "category": t.get("category", "Surveillance en continu"),
            "last_run": last_run.isoformat() if last_run else None,
            "next_run": next_run.isoformat() if next_run else None,
            "flux": None,
            "detail": t.get("detail"),
        })

    # Tâches par flux (flux_json_sources.json)
    for fname in ("flux_json_sources.json", "flux_json_sources.example.json"):
        flux_file = PROJECT_ROOT / "config" / fname
        if flux_file.exists():
            try:
                flux_list = json.loads(flux_file.read_text(encoding="utf-8"))
                for flux in flux_list:
                    title = flux.get("title", "Flux inconnu")
                    # Support format plat (cron) et format imbriqué (scheduler.cron)
                    cron = (flux.get("cron")
                            or flux.get("scheduler", {}).get("cron", "0 6 * * 1"))
                    flux_dir = PROJECT_ROOT / "data" / "articles" / title.strip().replace(" ", "-").replace("\u00a0", "-")
                    last_run = latest_mtime(flux_dir)
                    next_run = next_cron_occurrence(cron, now)
                    tasks.append({
                        "name": f"Flux : {title}",
                        "script": "Get_data_from_JSONFile_AskSummary_v2.py",
                        "cron": cron,
                        "label": cron_label(cron),
                        "last_run": last_run.isoformat() if last_run else None,
                        "next_run": next_run.isoformat() if next_run else None,
                        "flux": title,
                    })
            except (json.JSONDecodeError, KeyError, TypeError):
                pass
            break  # Utiliser uniquement le premier fichier de config existant

    return jsonify({"tasks": tasks, "now": now.isoformat()})


@app.route("/api/keywords", methods=["GET"])
def api_get_keywords():
    path = PROJECT_ROOT / "config" / "keyword-to-search.json"
    if not path.exists():
        return jsonify([])
    try:
        return jsonify(json.loads(path.read_text(encoding="utf-8")))
    except json.JSONDecodeError:
        return jsonify([])


@app.route("/api/keywords", methods=["POST"])
def api_save_keywords():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        abort(400, "Format invalide : tableau attendu")
    path = PROJECT_ROOT / "config" / "keyword-to-search.json"
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return jsonify({"ok": True})


@app.route("/api/rss-feeds", methods=["GET"])
def api_get_rss_feeds():
    """Parse le fichier OPML Reeder et retourne les flux RSS triés alphabétiquement."""
    import xml.etree.ElementTree as ET
    opml_path = PROJECT_ROOT / "data" / "WUDD.opml"
    if not opml_path.exists():
        return jsonify([])
    try:
        tree = ET.parse(opml_path)
        root = tree.getroot()
        feeds = []
        for o in root.findall(".//outline[@type='rss']"):
            title   = o.get("title") or o.get("text") or ""
            xml_url = o.get("xmlUrl") or ""
            html_url = o.get("htmlUrl") or ""
            if xml_url:
                feeds.append({"title": title, "xmlUrl": xml_url, "htmlUrl": html_url})
        feeds.sort(key=lambda f: f["title"].lower())
        return jsonify(feeds)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rss-feeds/check", methods=["POST"])
def api_check_rss_feed():
    """Vérifie si une URL RSS répond. Body JSON: {"url": "..."}"""
    import requests as req
    data = request.get_json(force=True) or {}
    url = data.get("url", "").strip()
    if not url:
        return jsonify({"ok": False, "error": "URL manquante"}), 400
    try:
        r = req.head(url, timeout=8, allow_redirects=True,
                     headers={"User-Agent": "WUDD.ai/1.0"})
        ok = r.status_code < 400
        return jsonify({"ok": ok, "status": r.status_code})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/rss-feeds/resolve", methods=["POST"])
def api_resolve_rss_feed():
    """Résout une URL RSS : vérifie qu'elle répond et extrait le titre du canal."""
    import requests as req
    import xml.etree.ElementTree as ET
    from urllib.parse import urlparse
    data = request.get_json(force=True) or {}
    url = data.get("url", "").strip()
    if not url or not url.startswith("http"):
        return jsonify({"ok": False, "error": "URL invalide"}), 400
    try:
        r = req.get(url, timeout=10, allow_redirects=True,
                    headers={"User-Agent": "WUDD.ai/1.0"})
        if r.status_code >= 400:
            return jsonify({"ok": False, "error": f"HTTP {r.status_code}"})
        title = ""
        html_url = ""
        try:
            root = ET.fromstring(r.content)
            # RSS 2.0
            chan = root.find("channel")
            if chan is not None:
                t = chan.find("title")
                if t is not None and t.text:
                    title = t.text.strip()
                lk = chan.find("link")
                if lk is not None and lk.text:
                    html_url = lk.text.strip()
            # Atom
            if not title:
                ns = {"atom": "http://www.w3.org/2005/Atom"}
                t = root.find("atom:title", ns) or root.find("title")
                if t is not None and t.text:
                    title = t.text.strip()
                lk = root.find("atom:link", ns)
                if lk is not None:
                    html_url = lk.get("href", "")
        except Exception:
            pass
        if not title:
            title = urlparse(url).netloc
        return jsonify({"ok": True, "title": title, "xmlUrl": url, "htmlUrl": html_url})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/rss-feeds/save", methods=["POST"])
def api_save_rss_feeds():
    """Sauvegarde la liste de flux dans data/WUDD.opml en respectant le format OPML."""
    import xml.etree.ElementTree as ET
    feeds = request.get_json(force=True)
    if not isinstance(feeds, list):
        return jsonify({"error": "Données invalides"}), 400
    opml_path = PROJECT_ROOT / "data" / "WUDD.opml"
    try:
        root = ET.Element("opml", version="2.0")
        head = ET.SubElement(root, "head")
        ET.SubElement(head, "title").text = "Reeder"
        body = ET.SubElement(root, "body")
        for f in feeds:
            ET.SubElement(body, "outline",
                          type="rss",
                          title=f.get("title", ""),
                          text=f.get("title", ""),
                          xmlUrl=f.get("xmlUrl", ""),
                          htmlUrl=f.get("htmlUrl", ""))
        tree = ET.ElementTree(root)
        ET.indent(tree, space="  ")
        with open(opml_path, "wb") as fh:
            tree.write(fh, encoding="UTF-8", xml_declaration=True)
        return jsonify({"ok": True, "count": len(feeds)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rss-feeds/stats", methods=["GET"])
def api_rss_feeds_stats():
    """Retourne le nombre d'articles et la date de dernière publication par domaine
    en scannant tous les fichiers JSON dans data/articles-from-rss/."""
    from urllib.parse import urlparse
    from email.utils import parsedate_to_datetime
    from datetime import datetime

    rss_dir = PROJECT_ROOT / "data" / "articles-from-rss"
    stats = {}  # domain -> {count, lastDate}

    if not rss_dir.exists():
        return jsonify({})

    for json_file in rss_dir.glob("*.json"):
        try:
            articles = json.loads(json_file.read_text(encoding="utf-8"))
            if not isinstance(articles, list):
                continue
            for article in articles:
                url = article.get("URL", "")
                date_str = article.get("Date de publication", "")
                if not url:
                    continue
                try:
                    hostname = urlparse(url).hostname or ""
                    domain = hostname.removeprefix("www.")
                except Exception:
                    continue
                if not domain:
                    continue
                entry = stats.setdefault(domain, {"count": 0, "lastDate": None})
                entry["count"] += 1
                if date_str:
                    dt = None
                    try:
                        dt = parsedate_to_datetime(date_str)
                    except Exception:
                        pass
                    if dt is None:
                        try:
                            dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                        except Exception:
                            pass
                    if dt is not None:
                        dt_iso = dt.isoformat()
                        if entry["lastDate"] is None or dt_iso > entry["lastDate"]:
                            entry["lastDate"] = dt_iso
        except Exception:
            continue

    return jsonify(stats)


@app.route("/api/web-sources", methods=["GET"])
def api_get_web_sources():
    """Retourne la liste des sources web depuis config/web_sources.json."""
    path = PROJECT_ROOT / "config" / "web_sources.json"
    if not path.exists():
        return jsonify([])
    try:
        return jsonify(json.loads(path.read_text(encoding="utf-8")))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/web-sources/save", methods=["POST"])
def api_save_web_sources():
    """Sauvegarde la liste des sources web dans config/web_sources.json."""
    sources = request.get_json(force=True)
    if not isinstance(sources, list):
        return jsonify({"error": "Données invalides"}), 400
    path = PROJECT_ROOT / "config" / "web_sources.json"
    try:
        tmp = path.with_suffix(".tmp")
        tmp.write_text(json.dumps(sources, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(path)
        return jsonify({"ok": True, "count": len(sources)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/web-sources/check", methods=["POST"])
def api_check_web_source():
    """Vérifie si une URL de sitemap ou de site est accessible. Body JSON: {"url": "..."}"""
    import requests as req
    data = request.get_json(force=True) or {}
    url = data.get("url", "").strip()
    if not url:
        return jsonify({"ok": False, "error": "URL manquante"}), 400
    try:
        r = req.head(url, timeout=8, allow_redirects=True,
                     headers={"User-Agent": "WUDD.ai/2.2"})
        ok = r.status_code < 400
        return jsonify({"ok": ok, "status": r.status_code})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/web-sources/resolve", methods=["POST"])
def api_resolve_web_source():
    """Résout une URL de site web : extrait le titre et détecte le sitemap.

    Body JSON: {"url": "https://example.com"}
    Retourne: {ok, title, base_url, sitemap_url, html_url}
    """
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse, urljoin

    data = request.get_json(force=True) or {}
    url = data.get("url", "").strip()
    if not url or not url.startswith("http"):
        return jsonify({"ok": False, "error": "URL invalide"}), 400

    headers = {"User-Agent": "Mozilla/5.0 (compatible; WUDD.ai/2.2)"}
    try:
        r = req.get(url, timeout=10, allow_redirects=True, headers=headers)
        if r.status_code >= 400:
            return jsonify({"ok": False, "error": f"HTTP {r.status_code}"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

    soup = BeautifulSoup(r.content, "html.parser")
    parsed = urlparse(r.url)
    base_url = f"{parsed.scheme}://{parsed.netloc}"

    # Titre du site
    title = ""
    og = soup.find("meta", property="og:site_name")
    if og:
        title = og.get("content", "").strip()
    if not title:
        og = soup.find("meta", property="og:title")
        if og:
            title = og.get("content", "").strip()
    if not title:
        t = soup.find("title")
        if t:
            title = t.get_text(strip=True)
    if not title:
        title = parsed.netloc.replace("www.", "")

    # Détection du sitemap
    sitemap_url = ""
    # 1. Balise <link rel="sitemap">
    link_tag = soup.find("link", rel=lambda v: v and "sitemap" in (v if isinstance(v, str) else " ".join(v)).lower())
    if link_tag:
        sitemap_url = urljoin(base_url, link_tag.get("href", ""))

    # 2. Essai /sitemap.xml
    if not sitemap_url:
        candidates = ["/sitemap.xml", "/sitemap_index.xml", "/sitemap.xml.gz"]
        for cand in candidates:
            try:
                test_url = base_url + cand
                tr = req.head(test_url, timeout=5, headers=headers, allow_redirects=True)
                if tr.status_code < 400:
                    sitemap_url = test_url
                    break
            except Exception:
                continue

    return jsonify({
        "ok": True,
        "title": title,
        "base_url": base_url,
        "sitemap_url": sitemap_url,
        "html_url": r.url,
    })


@app.route("/api/web-sources/state", methods=["GET"])
def api_web_sources_state():
    """Retourne l'état du web_watcher : nombre d'URLs traitées par source."""
    state_path = PROJECT_ROOT / "data" / "web_watcher_state.json"
    if not state_path.exists():
        return jsonify({})
    try:
        state = json.loads(state_path.read_text(encoding="utf-8"))
        summary = {
            name: len(v.get("processed_urls", []))
            for name, v in state.items()
        }
        return jsonify(summary)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/flux-sources", methods=["GET"])
def api_get_flux_sources():
    path = PROJECT_ROOT / "config" / "flux_json_sources.json"
    if not path.exists():
        # Retourner le fichier exemple si disponible
        example = PROJECT_ROOT / "config" / "flux_json_sources.example.json"
        if example.exists():
            try:
                return jsonify(json.loads(example.read_text(encoding="utf-8")))
            except json.JSONDecodeError:
                pass
        return jsonify([])
    try:
        return jsonify(json.loads(path.read_text(encoding="utf-8")))
    except json.JSONDecodeError:
        return jsonify([])


@app.route("/api/flux-sources", methods=["POST"])
def api_save_flux_sources():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        abort(400, "Format invalide : tableau attendu")
    path = PROJECT_ROOT / "config" / "flux_json_sources.json"
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return jsonify({"ok": True})


# ─── Quota ────────────────────────────────────────────────────────────────────

@app.route("/api/quota/config", methods=["GET"])
def api_get_quota_config():
    """Retourne la configuration des quotas (config/quota.json)."""
    from utils.quota import get_quota_manager, DEFAULT_CONFIG
    mgr = get_quota_manager()
    # Retourner la config avec les valeurs par défaut en fallback
    cfg = {**DEFAULT_CONFIG, **mgr._config}
    return jsonify(cfg)


@app.route("/api/quota/config", methods=["POST"])
def api_save_quota_config():
    """Sauvegarde la configuration des quotas."""
    from utils.quota import get_quota_manager
    data = request.get_json(force=True)
    if not isinstance(data, dict):
        abort(400, "Format invalide : objet attendu")
    # Validation basique des types
    for int_key in ("global_daily_limit", "per_keyword_daily_limit", "per_source_daily_limit",
                    "per_entity_daily_limit", "summary_max_lines"):
        if int_key in data:
            try:
                data[int_key] = max(1, int(data[int_key]))
            except (ValueError, TypeError):
                abort(400, f"Valeur invalide pour {int_key}")
    get_quota_manager().save_config(data)
    # Invalider le singleton Config pour que summary_max_lines soit rechargé
    try:
        from utils.config import get_config as _get_config
        _get_config(force_reload=True)
    except Exception:
        pass
    return jsonify({"ok": True})


@app.route("/api/quota/stats", methods=["GET"])
def api_get_quota_stats():
    """Retourne les statistiques de consommation du jour."""
    from utils.quota import get_quota_manager
    return jsonify(get_quota_manager().get_stats())


@app.route("/api/quota/reset", methods=["POST"])
def api_reset_quota():
    """Réinitialise les compteurs de quota du jour."""
    from utils.quota import get_quota_manager
    get_quota_manager().reset_day()
    return jsonify({"ok": True})


@app.route("/api/search/entity")
def api_search_entity():
    """Recherche cross-fichiers d'une valeur d'entité nommée."""
    q = request.args.get("q", "").strip()
    entity_type = request.args.get("type", "").strip()
    if len(q) < 2:
        return jsonify([])

    q_lower = q.lower()
    results = []

    data_dirs = [
        PROJECT_ROOT / "data" / "articles",
        PROJECT_ROOT / "data" / "articles-from-rss",
    ]

    for data_dir in data_dirs:
        if not data_dir.exists():
            continue
        for json_file in sorted(data_dir.rglob("*.json")):
            if "cache" in json_file.relative_to(data_dir).parts:
                continue
            try:
                articles = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if not isinstance(articles, list):
                    continue
            except (json.JSONDecodeError, OSError):
                continue

            for article in articles:
                ents = article.get("entities")
                if not ents or not isinstance(ents, dict):
                    continue

                matched_types = []
                for etype, values in ents.items():
                    if entity_type and etype != entity_type:
                        continue
                    if not isinstance(values, list):
                        continue
                    if any(q_lower in str(v).lower() for v in values):
                        matched_types.append(etype)

                if not matched_types:
                    continue

                resume = article.get("Résumé", "")
                idx = resume.lower().find(q_lower)
                if idx >= 0:
                    start = max(0, idx - 80)
                    end = min(len(resume), idx + len(q) + 80)
                    excerpt = (
                        ("…" if start > 0 else "")
                        + resume[start:end]
                        + ("…" if end < len(resume) else "")
                    )
                else:
                    excerpt = resume[:160] + ("…" if len(resume) > 160 else "")

                rel = json_file.relative_to(PROJECT_ROOT)
                results.append({
                    "path": str(rel).replace("\\", "/"),
                    "name": json_file.name,
                    "source": article.get("Sources", ""),
                    "date": article.get("Date de publication", ""),
                    "url": article.get("URL", ""),
                    "excerpt": excerpt,
                    "types": matched_types,
                })

    results.sort(key=lambda r: r["date"], reverse=True)
    return jsonify(results[:100])


@app.route("/api/entities/dashboard")
def api_entities_dashboard():
    """Agrège les entités de tous les fichiers JSON et retourne des stats globales."""
    data_dirs = [
        PROJECT_ROOT / "data" / "articles",
        PROJECT_ROOT / "data" / "articles-from-rss",
    ]

    total_files = 0
    total_articles = 0
    total_with_entities = 0
    # { type: { value: count } }
    by_type: dict[str, dict[str, int]] = {}

    for data_dir in data_dirs:
        if not data_dir.exists():
            continue
        for json_file in sorted(data_dir.rglob("*.json")):
            if "cache" in json_file.relative_to(data_dir).parts:
                continue
            try:
                articles = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if not isinstance(articles, list):
                    continue
            except (json.JSONDecodeError, OSError):
                continue

            total_files += 1
            total_articles += len(articles)

            for article in articles:
                ents = article.get("entities")
                if not ents or not isinstance(ents, dict):
                    continue
                has_ent = False
                for etype, values in ents.items():
                    if not isinstance(values, list) or not values:
                        continue
                    has_ent = True
                    if etype not in by_type:
                        by_type[etype] = {}
                    for v in values:
                        if isinstance(v, str) and v.strip():
                            key = v.strip()
                            by_type[etype][key] = by_type[etype].get(key, 0) + 1
                if has_ent:
                    total_with_entities += 1

    result_types = []
    for etype, value_counts in by_type.items():
        sorted_values = sorted(value_counts.items(), key=lambda x: x[1], reverse=True)
        result_types.append({
            "type": etype,
            "unique_count": len(sorted_values),
            "mention_count": sum(c for _, c in sorted_values),
            "top": [{"value": v, "count": c} for v, c in sorted_values[:50]],
        })
    result_types.sort(key=lambda x: x["mention_count"], reverse=True)

    return jsonify({
        "total_files": total_files,
        "total_articles": total_articles,
        "total_with_entities": total_with_entities,
        "by_type": result_types,
    })


@app.route("/api/entities/articles")
def api_entities_articles():
    """Retourne tous les articles contenant une entité donnée (type + valeur)."""
    entity_type = request.args.get("type", "").strip()
    entity_value = request.args.get("value", "").strip()
    if not entity_type or not entity_value:
        return jsonify({"error": "Paramètres type et value requis"}), 400

    seen_urls = set()
    results = []
    for data_dir in [PROJECT_ROOT / "data" / "articles", PROJECT_ROOT / "data" / "articles-from-rss"]:
        if not data_dir.exists():
            continue
        for json_file in sorted(data_dir.rglob("*.json")):
            if "cache" in json_file.relative_to(data_dir).parts:
                continue
            try:
                articles = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if not isinstance(articles, list):
                    continue
            except (json.JSONDecodeError, OSError):
                continue
            for article in articles:
                entities = article.get("entities", {})
                if not isinstance(entities, dict):
                    continue
                values = entities.get(entity_type, [])
                if not (isinstance(values, list) and entity_value in values):
                    continue
                # Déduplication : URL d'abord, puis résumé (articles syndiqués sans URL unique)
                url = (article.get("URL") or "").strip()
                resume_key = article.get("Résumé", "")[:150].strip()
                if (url and url in seen_urls) or (resume_key and resume_key in seen_urls):
                    continue
                if url:
                    seen_urls.add(url)
                if resume_key:
                    seen_urls.add(resume_key)
                results.append(article)

    results.sort(key=lambda a: a.get("Date de publication", ""), reverse=True)
    return jsonify(results)


@app.route("/api/entity-context")
def api_entity_context():
    """Construit un bloc de contexte structuré pour une entité donnée.

    Utilisé par le Terminal IA pour pré-charger toutes les informations
    disponibles sur une entité : articles, co-occurrences, calendrier.

    Query params :
      type  — type NER (ex. "ORG", "PERSON", "GPE")
      value — valeur de l'entité (ex. "OpenAI", "Emmanuel Macron")
      n     — nombre max d'articles à inclure dans le contexte (défaut 20)
    """
    entity_type  = request.args.get("type",  "").strip()
    entity_value = request.args.get("value", "").strip()
    n_articles   = min(int(request.args.get("n", 20)), 50)

    if not entity_type or not entity_value:
        return jsonify({"error": "Paramètres type et value requis"}), 400

    # ── 1. Articles ────────────────────────────────────────────────────────────
    seen_urls: set = set()
    articles: list[dict] = []
    for data_dir in [PROJECT_ROOT / "data" / "articles",
                     PROJECT_ROOT / "data" / "articles-from-rss"]:
        if not data_dir.exists():
            continue
        for json_file in sorted(data_dir.rglob("*.json")):
            if "cache" in json_file.relative_to(data_dir).parts:
                continue
            try:
                arts = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if not isinstance(arts, list):
                    continue
            except (json.JSONDecodeError, OSError):
                continue
            for article in arts:
                entities = article.get("entities", {})
                if not isinstance(entities, dict):
                    continue
                values = entities.get(entity_type, [])
                if not (isinstance(values, list) and entity_value in values):
                    continue
                url = (article.get("URL") or "").strip()
                resume_key = article.get("Résumé", "")[:150].strip()
                if (url and url in seen_urls) or (resume_key and resume_key in seen_urls):
                    continue
                if url:
                    seen_urls.add(url)
                if resume_key:
                    seen_urls.add(resume_key)
                articles.append(article)

    articles.sort(key=lambda a: a.get("Date de publication", ""), reverse=True)
    top_articles = articles[:n_articles]

    # ── 2. Co-occurrences L1 ──────────────────────────────────────────────────
    from collections import Counter as _Counter, defaultdict as _defaultdict
    cooc: _Counter = _Counter()
    for article in articles:
        ents = article.get("entities", {})
        if not isinstance(ents, dict):
            continue
        others: list[tuple[str, str]] = []
        for etype, evals in ents.items():
            if not isinstance(evals, list):
                continue
            for ev in evals:
                if not (etype == entity_type and ev == entity_value):
                    others.append((etype, ev))
        for pair in others:
            cooc[pair] += 1

    top_cooc = cooc.most_common(15)

    # ── 3. Calendrier (fréquence mensuelle) ───────────────────────────────────
    monthly: dict[str, int] = _defaultdict(int)
    for article in articles:
        date_str = article.get("Date de publication", "")
        if date_str and len(date_str) >= 7:
            # "JJ/MM/AAAA" → "AAAA-MM" ou "AAAA-MM-JJ" ISO → "AAAA-MM"
            if "/" in date_str:
                parts = date_str.split("/")
                if len(parts) == 3:
                    monthly[f"{parts[2]}-{parts[1]}"] += 1
            elif "-" in date_str:
                monthly[date_str[:7]] += 1

    calendar_lines = [f"  {month} : {count} article(s)"
                      for month, count in sorted(monthly.items(), reverse=True)[:12]]

    # ── 4. Construction du bloc texte ─────────────────────────────────────────
    type_labels = {
        "PERSON": "Personne", "ORG": "Organisation", "GPE": "Pays/Région",
        "LOC": "Lieu", "PRODUCT": "Produit", "EVENT": "Événement",
        "DATE": "Date", "MONEY": "Montant",
    }
    type_label = type_labels.get(entity_type, entity_type)

    lines: list[str] = [
        f"# Contexte entité : {entity_value} ({type_label})",
        f"Total articles trouvés : {len(articles)}",
        "",
    ]

    # Calendrier
    if calendar_lines:
        lines.append("## Calendrier des mentions (derniers 12 mois)")
        lines.extend(calendar_lines)
        lines.append("")

    # Co-occurrences
    if top_cooc:
        lines.append("## Entités co-occurrentes principales")
        for (etype, ev), count in top_cooc:
            lbl = type_labels.get(etype, etype)
            lines.append(f"  - {ev} ({lbl}) : {count} co-occurrence(s)")
        lines.append("")

    # Sentiments agrégés
    sentiments: _Counter = _Counter()
    for art in articles:
        s = art.get("sentiment")
        if s:
            sentiments[s] += 1
    if sentiments:
        lines.append("## Tonalité éditoriale")
        for sent, cnt in sentiments.most_common():
            lines.append(f"  - {sent} : {cnt} article(s)")
        lines.append("")

    # Sources
    sources: _Counter = _Counter()
    for art in articles:
        src = art.get("Sources")
        if src:
            sources[src] += 1
    if sources:
        lines.append("## Sources principales")
        for src, cnt in sources.most_common(8):
            lines.append(f"  - {src} : {cnt} article(s)")
        lines.append("")

    # Articles (résumés tronqués)
    if top_articles:
        lines.append(f"## Articles récents ({len(top_articles)} sur {len(articles)})")
        for i, art in enumerate(top_articles, 1):
            date  = art.get("Date de publication", "?")
            src   = art.get("Sources", "?")
            url   = art.get("URL", "")
            resume = (art.get("Résumé") or "").strip()
            if len(resume) > 500:
                resume = resume[:500] + "…"
            header = f"### {i}. [{date}] {src}"
            if url:
                header += f" — {url}"
            lines.append(header)
            if resume:
                lines.append(resume)
            lines.append("")

    context_text = "\n".join(lines)

    return jsonify({
        "entity_type":   entity_type,
        "entity_value":  entity_value,
        "article_count": len(articles),
        "context_text":  context_text,
    })


@app.route("/api/entities/cooccurrences")
def api_entities_cooccurrences():
    """Retourne les entités co-occurrentes d'une entité donnée (via articles partagés).

    Paramètres :
      type, value  — entité centrale
      limit        — max d'entités niveau 1 (défaut 40)
      depth        — profondeur du graphe : 1 ou 2 (défaut 1)
      limit_l2     — max d'entités niveau 2 par nœud L1 (défaut 4)
    """
    entity_type = request.args.get("type", "").strip()
    entity_value = request.args.get("value", "").strip()
    depth = min(int(request.args.get("depth", 1)), 2)
    # Quand depth=2 on réduit L1 pour garder le graphe lisible
    limit_l1 = min(int(request.args.get("limit", 40)), 100)
    if depth >= 2:
        limit_l1 = min(limit_l1, 12)
    limit_l2 = min(int(request.args.get("limit_l2", 4)), 15)

    if not entity_type or not entity_value:
        return jsonify({"error": "Paramètres type et value requis"}), 400

    def node_id(t, v):
        return f"{t}:{v}"

    # ── Chargement unique de tous les articles ────────────────────────────────
    all_articles: list[dict] = []
    for data_dir in [PROJECT_ROOT / "data" / "articles", PROJECT_ROOT / "data" / "articles-from-rss"]:
        if not data_dir.exists():
            continue
        for json_file in sorted(data_dir.rglob("*.json")):
            if "cache" in json_file.relative_to(data_dir).parts:
                continue
            try:
                arts = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if isinstance(arts, list):
                    all_articles.extend(arts)
            except (json.JSONDecodeError, OSError):
                continue

    # ── Passe 1 : co-occurrences L1 ──────────────────────────────────────────
    cooc_l1: dict[tuple[str, str], int] = {}
    for article in all_articles:
        entities = article.get("entities", {})
        if not isinstance(entities, dict):
            continue
        values = entities.get(entity_type, [])
        if not isinstance(values, list) or entity_value not in values:
            continue
        for etype, evals in entities.items():
            if not isinstance(evals, list):
                continue
            for ev in evals:
                if etype == entity_type and ev == entity_value:
                    continue
                key = (etype, ev)
                cooc_l1[key] = cooc_l1.get(key, 0) + 1

    sorted_l1 = sorted(cooc_l1.items(), key=lambda x: x[1], reverse=True)[:limit_l1]
    top_l1_set: set[tuple[str, str]] = {k for k, _ in sorted_l1}

    # ── Construction des nœuds / arêtes L1 ───────────────────────────────────
    nodes = [{"type": entity_type, "value": entity_value, "count": 0,
               "central": True, "level": 0}]
    edges = []

    for (etype, ev), count in sorted_l1:
        nodes.append({"type": etype, "value": ev, "count": count,
                       "central": False, "level": 1})
        edges.append({"source": node_id(entity_type, entity_value),
                       "target": node_id(etype, ev), "weight": count})

    # ── Passe 2 : co-occurrences L2 (optionnel) ──────────────────────────────
    if depth >= 2 and top_l1_set:
        # Pour chaque article, identifie les entités L1 présentes, puis
        # accumule leurs co-occurrences (→ candidats L2).
        cooc_l2: dict[tuple[tuple, tuple], int] = {}
        for article in all_articles:
            entities = article.get("entities", {})
            if not isinstance(entities, dict):
                continue
            # Entités L1 présentes dans cet article
            l1_here = set()
            for etype, evals in entities.items():
                if not isinstance(evals, list):
                    continue
                for ev in evals:
                    if (etype, ev) in top_l1_set:
                        l1_here.add((etype, ev))
            if not l1_here:
                continue
            # Co-occurrences entre chaque nœud L1 et les autres entités
            for l1_key in l1_here:
                for etype, evals in entities.items():
                    if not isinstance(evals, list):
                        continue
                    for ev in evals:
                        co_key = (etype, ev)
                        if co_key == l1_key:
                            continue
                        if co_key == (entity_type, entity_value):
                            continue  # évite l'arête de retour vers le centre
                        cooc_l2[(l1_key, co_key)] = cooc_l2.get((l1_key, co_key), 0) + 1

        # Regroupe par nœud L1
        l1_coocs: dict[tuple, list] = {}
        for (l1_key, co_key), count in cooc_l2.items():
            l1_coocs.setdefault(l1_key, []).append((co_key, count))

        existing: set[tuple[str, str]] = {(entity_type, entity_value)} | top_l1_set
        added_l2: set[tuple[str, str]] = set()

        for l1_key, coocs in l1_coocs.items():
            top_for_l1 = sorted(
                [x for x in coocs if x[0] not in existing],
                key=lambda x: x[1],
                reverse=True,
            )[:limit_l2]
            for (etype, ev), count in top_for_l1:
                l2_key = (etype, ev)
                if l2_key not in added_l2:
                    nodes.append({"type": etype, "value": ev, "count": count,
                                   "central": False, "level": 2})
                    added_l2.add(l2_key)
                    existing.add(l2_key)
                edges.append({"source": node_id(*l1_key),
                               "target": node_id(etype, ev),
                               "weight": count})

    return jsonify({"nodes": nodes, "edges": edges, "total_cooc": len(cooc_l1)})


@app.route("/api/entities/geocode", methods=["POST"])
def api_entities_geocode():
    """Géocode une liste d'entités via Wikipedia API avec cache JSON local."""
    import requests as req

    names = request.get_json(force=True) or []
    if not names or not isinstance(names, list):
        return jsonify({})

    cache_path = PROJECT_ROOT / "data" / "geocode_cache.json"
    cache = {}
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            cache = {}

    to_fetch = [n for n in names if n not in cache]

    WIKIPEDIA_UA = (
        "WUDD.ai/2.1.0 (news monitoring tool; "
        "https://github.com/patrickostertag) python-requests"
    )

    BATCH = 50
    for i in range(0, len(to_fetch), BATCH):
        batch = to_fetch[i : i + BATCH]
        titles_str = "|".join(batch)
        fetched_coords: dict[str, dict] = {}

        for lang in ("fr", "en"):
            try:
                r = req.get(
                    f"https://{lang}.wikipedia.org/w/api.php",
                    params={
                        "action": "query",
                        "titles": titles_str,
                        "prop": "coordinates",
                        "format": "json",
                        "origin": "*",
                    },
                    headers={"User-Agent": WIKIPEDIA_UA},
                    timeout=10,
                )
                data = r.json()
                pages = data.get("query", {}).get("pages", {})
                # normalizations : associe les redirections aux noms originaux
                normalized = {
                    n["from"]: n["to"]
                    for n in data.get("query", {}).get("normalized", [])
                }
                for page in pages.values():
                    if "coordinates" not in page:
                        continue
                    title = page.get("title", "")
                    coords = {
                        "lat": page["coordinates"][0]["lat"],
                        "lon": page["coordinates"][0]["lon"],
                    }
                    fetched_coords[title] = coords
                    # mappe aussi la forme originale si redirigée
                    for orig, norm in normalized.items():
                        if norm == title:
                            fetched_coords[orig] = coords
            except Exception:
                continue

            if lang == "fr" and len(fetched_coords) >= len(batch):
                break  # tout trouvé en FR

        for name in batch:
            if name not in cache:
                cache[name] = fetched_coords.get(name)  # None si introuvable

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    return jsonify({n: cache.get(n) for n in names})


@app.route("/api/entities/images", methods=["POST"])
def api_entities_images():
    """Récupère les images Wikipedia d'une liste d'entités.

    Accepte [{name, type}] ou [str] (compat. ascendante).
    Stratégie :
      - PERSON               → Wikipedia pageimages (portrait)
      - ORG / PRODUCT        → Wikidata P154 (logo officiel) + fallback pageimages
      - autres / inconnus    → Wikipedia pageimages
    """
    import requests as req

    body = request.get_json(force=True) or []
    if not body or not isinstance(body, list):
        return jsonify({})

    # Normalise l'entrée en [{name, type}]
    entities: list[dict] = []
    for item in body:
        if isinstance(item, dict):
            entities.append({"name": item.get("name", "").strip(), "type": item.get("type", "").upper()})
        elif isinstance(item, str):
            entities.append({"name": item.strip(), "type": ""})
    entities = [e for e in entities if e["name"]]

    UA = "WUDD.ai/2.1.0 (news monitoring tool; https://github.com/patrickostertag) python-requests"
    THUMB = 200
    BATCH = 50

    cache_path = PROJECT_ROOT / "data" / "images_cache.json"
    cache: dict = {}
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            cache = {}

    to_fetch = [e for e in entities if e["name"] not in cache]
    if not to_fetch:
        return jsonify({e["name"]: cache.get(e["name"]) for e in entities})

    # ── Séparer PERSON vs ORG/PRODUCT vs autres ──────────────────────────────
    person_names = [e["name"] for e in to_fetch if e["type"] == "PERSON"]
    logo_names   = [e["name"] for e in to_fetch if e["type"] in ("ORG", "PRODUCT")]
    other_names  = [e["name"] for e in to_fetch if e["type"] not in ("PERSON", "ORG", "PRODUCT")]

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _pageimages(names_batch: list[str]) -> dict[str, dict]:
        """Retourne {name: {url,width,height}} via Wikipedia pageimages."""
        result: dict[str, dict] = {}
        for i in range(0, len(names_batch), BATCH):
            batch = names_batch[i : i + BATCH]
            titles_str = "|".join(batch)
            for lang in ("fr", "en"):
                try:
                    r = req.get(
                        f"https://{lang}.wikipedia.org/w/api.php",
                        params={"action": "query", "titles": titles_str,
                                "prop": "pageimages", "pithumbsize": THUMB,
                                "pilicense": "any", "format": "json", "origin": "*"},
                        headers={"User-Agent": UA}, timeout=10,
                    )
                    pages = r.json().get("query", {})
                    normalized = {n["from"]: n["to"] for n in pages.get("normalized", [])}
                    for page in pages.get("pages", {}).values():
                        if "thumbnail" not in page:
                            continue
                        title = page["title"]
                        img = {"url": page["thumbnail"]["source"],
                               "width": page["thumbnail"].get("width", THUMB),
                               "height": page["thumbnail"].get("height", THUMB)}
                        result[title] = img
                        for orig, norm in normalized.items():
                            if norm == title:
                                result[orig] = img
                except Exception:
                    continue
                if lang == "fr" and len(result) >= len(batch):
                    break
        return result

    def _wikidata_logos(names_batch: list[str]) -> tuple[dict[str, str], set[str]]:
        """Retourne ({name: image_filename}, rejected) via Wikidata P154 puis P18.

        rejected : entités dont l'article Wikipedia correspond à une personne,
        un prénom, un concept sans lien avec une ORG/PRODUCT → pas de fallback.

        Règle :
          - P31 ∈ HARD_WRONG (humain, prénom, template) → rejet définitif
          - P31 ∈ SOFT_WRONG (homonymie, liste) → pas de logo Wikidata, fallback search autorisé
          - P154 (logo officiel) → prioritaire
          - P18 (image générale) → fallback si pas de P154 ET P31 ∈ OK_TYPES
          - P31 ∈ OK_TYPES mais pas P18 → fallback _search_pageimage_single
          - P31 absent ou inconnu sans P18 → fallback _search_pageimage_single
        """
        logos: dict[str, str] = {}
        rejected: set[str] = set()
        P154 = "P154"
        P18  = "P18"

        # Hard-rejects : entité DÉFINITIVEMENT hors-scope ORG/PRODUCT → cache=None, pas de fallback
        HARD_WRONG = {
            "Q5",        # human / personne
            "Q202444",   # given name / prénom
            "Q101352",   # family name / nom de famille
            "Q11266439", # Wikimedia template
        }
        # Soft-rejects : titre ambigu ou liste → pas de logo Wikidata, mais fallback search autorisé
        SOFT_WRONG = {
            "Q4167410",  # Wikimedia disambiguation page
            "Q50339617", # Wikimedia list article
        }
        WRONG_TYPES = HARD_WRONG | SOFT_WRONG  # union pour le check P31
        # Types P31 compatibles avec une entité ORG/PRODUCT (autorise P18 comme image)
        OK_TYPES = {
            "Q4830453", "Q783794", "Q891723", "Q43229", "Q167037",  # entreprises/orgs
            "Q7397", "Q166142", "Q9143", "Q9135", "Q7889",           # logiciels/tech
            "Q18127206", "Q18662854", "Q1331793", "Q17155032",       # tech/média
            "Q3220391", "Q122759350", "Q6576792", "Q118140435",      # réseaux sociaux / plateformes
        }

        def _filename(claim_value) -> str:
            """Extrait le nom de fichier d'un claim Wikidata (string ou dict)."""
            if isinstance(claim_value, str):
                return claim_value
            return claim_value.get("value", "") if isinstance(claim_value, dict) else ""

        for i in range(0, len(names_batch), BATCH):
            batch = names_batch[i : i + BATCH]
            titles_str = "|".join(batch)
            for site in ("enwiki", "frwiki"):
                try:
                    r = req.get(
                        "https://www.wikidata.org/w/api.php",
                        params={"action": "wbgetentities", "sites": site,
                                "titles": titles_str, "props": "claims|sitelinks",
                                "format": "json", "origin": "*"},
                        headers={"User-Agent": UA}, timeout=10,
                    )
                    for eid, entity in r.json().get("entities", {}).items():
                        if eid.startswith("-"):
                            continue
                        wiki_title = entity.get("sitelinks", {}).get(site, {}).get("title", "")
                        claims = entity.get("claims", {})
                        p31_ids = {
                            claim["mainsnak"]["datavalue"]["value"]["id"]
                            for claim in claims.get("P31", [])
                            if claim["mainsnak"].get("datavalue")
                        }
                        for orig in batch:
                            if orig.lower() == wiki_title.lower() and orig not in logos and orig not in rejected:
                                if p31_ids & HARD_WRONG:
                                    # Personne, prénom, template → rejet définitif (cache=None, pas de fallback)
                                    rejected.add(orig)
                                elif p31_ids & SOFT_WRONG:
                                    # Page de désambiguïsation ou liste → pas de logo Wikidata,
                                    # mais le fallback search peut quand même essayer
                                    break
                                elif P154 in claims:
                                    # Logo officiel (prioritaire)
                                    logos[orig] = _filename(claims[P154][0]["mainsnak"]["datavalue"]["value"])
                                elif P18 in claims and p31_ids & OK_TYPES:
                                    # Image générale Wikidata uniquement si type ORG/PRODUCT confirmé
                                    # (évite de retourner une image de forme géométrique, ex. Pentagone)
                                    logos[orig] = _filename(claims[P18][0]["mainsnak"]["datavalue"]["value"])
                                elif p31_ids & OK_TYPES:
                                    # Type ORG/PRODUCT confirmé mais sans image → fallback _search_pageimage_single
                                    pass
                                else:
                                    # P31 absent ou type inconnu → pas de logo Wikidata,
                                    # laisse _search_pageimage_single décider (cherche en anglais + valide type)
                                    break
                                break
                except Exception:
                    continue
        return logos, rejected

    def _wikidata_p18_persons(names_batch: list[str]) -> dict[str, str]:
        """Retourne {name: image_filename} via Wikidata P18 pour les PERSON.

        Utilisé en fallback quand Wikipedia pageimages ne trouve pas de portrait.
        Contrairement à _wikidata_logos, aucun filtre de type (P31 ignoré).
        """
        logos: dict[str, str] = {}
        P18 = "P18"
        for i in range(0, len(names_batch), BATCH):
            batch = names_batch[i : i + BATCH]
            titles_str = "|".join(batch)
            for site in ("enwiki", "frwiki"):
                try:
                    r = req.get(
                        "https://www.wikidata.org/w/api.php",
                        params={"action": "wbgetentities", "sites": site,
                                "titles": titles_str, "props": "claims|sitelinks",
                                "format": "json", "origin": "*"},
                        headers={"User-Agent": UA}, timeout=10,
                    )
                    for eid, entity in r.json().get("entities", {}).items():
                        if eid.startswith("-"):
                            continue
                        wiki_title = entity.get("sitelinks", {}).get(site, {}).get("title", "")
                        claims = entity.get("claims", {})
                        if P18 not in claims:
                            continue
                        for orig in batch:
                            if orig.lower() == wiki_title.lower() and orig not in logos:
                                val = claims[P18][0]["mainsnak"]["datavalue"]["value"]
                                fname = val if isinstance(val, str) else val.get("value", "")
                                if fname:
                                    logos[orig] = fname
                                break
                except Exception:
                    continue
        return logos

    def _resolve_logo_urls(filenames: list[str]) -> dict[str, str]:
        """Retourne {filename: url_miniature} depuis Wikimedia Commons."""
        urls: dict[str, str] = {}
        for i in range(0, len(filenames), BATCH):
            batch = filenames[i : i + BATCH]
            titles = "|".join(f"File:{f}" for f in batch)
            try:
                r = req.get(
                    "https://commons.wikimedia.org/w/api.php",
                    params={"action": "query", "titles": titles,
                            "prop": "imageinfo", "iiprop": "url",
                            "iiurlwidth": THUMB, "format": "json", "origin": "*"},
                    headers={"User-Agent": UA}, timeout=10,
                )
                for page in r.json().get("query", {}).get("pages", {}).values():
                    fname = page.get("title", "").removeprefix("File:")
                    info = page.get("imageinfo", [])
                    if info:
                        url = info[0].get("thumburl") or info[0].get("url")
                        if url:
                            urls[fname] = url
            except Exception:
                pass
        return urls

    # Types Wikidata à rejeter pour les entités ORG/PRODUCT dans le fallback search
    SEARCH_WRONG = {
        "Q5",        # human / personne physique
        "Q202444",   # given name / prénom
        "Q101352",   # family name / nom de famille
        "Q4167410",  # disambiguation page
        "Q11266439", # Wikimedia template
        "Q50339617", # Wikimedia list article
        "Q4086834",  # polygon (géométrie)
        "Q35234",    # regular polygon
        "Q12503",    # pentagon (forme géométrique)
        "Q8091",     # geometry (discipline)
        "Q1298765",  # mathematical object
        "Q17451",    # typeface (ex. Blackboard bold → article sur la lettre X)
        "Q58481926", # typeface family (ex. Fraktur)
        # Concepts linguistiques — ex. "Word" (PRODUCT) → article "Mot" (Q8171)
        "Q8171",     # word / mot (unité linguistique)
        "Q58778",    # linguistic concept (concept linguistique)
        "Q82042",    # part of speech (partie du discours)
        # Taxons biologiques — ex. "Apple" (ORG) → article "Pomme / Malus domestica"
        "Q16521",    # taxon (espèce végétale, animale…)
        "Q89",       # Malus domestica (pommier / pomme)
        "Q1364",     # apple (fruit générique)
    }

    def _wikidata_type_ok(qid: str, strict: bool = False) -> bool:
        """Retourne True si le QID Wikidata n'est pas un SEARCH_WRONG (ORG/PRODUCT check).

        strict=True : exige P31 non-vide (rejette les entités sans type défini,
        ex. lettre 'X' Q9968 qui n'a pas de P31 mais n'est pas une ORG).
        """
        if not qid:
            return not strict  # strict=True rejette les entités sans QID
        try:
            r2 = req.get(
                "https://www.wikidata.org/w/api.php",
                params={"action": "wbgetentities", "ids": qid,
                        "props": "claims", "format": "json", "origin": "*"},
                headers={"User-Agent": UA}, timeout=5,
            )
            claims = r2.json().get("entities", {}).get(qid, {}).get("claims", {})
            p31_ids = {
                c["mainsnak"]["datavalue"]["value"]["id"]
                for c in claims.get("P31", [])
                if c["mainsnak"].get("datavalue")
            }
            if strict and not p31_ids:
                return False  # pas de P31 → type inconnu, rejeté en mode strict
            return not bool(p31_ids & SEARCH_WRONG)
        except Exception:
            return True  # erreur réseau → on ne bloque pas

    def _search_pageimage_single(name: str, entity_type: str = "") -> tuple[str, dict | None]:
        """Fallback final : recherche Wikipedia generator=search avec validation de type.

        - ORG/PRODUCT : cherche en français d'abord (gsrlimit=3 pour parcourir les
          résultats ambigus, ex. 'Pentagone' → 1er résultat géométrique ignoré,
          2e résultat = Pentagone (États-Unis) retenu).
          Valide le type Wikidata de chaque résultat pour exclure formes géom., etc.
        - PERSON / autres : cherche en français d'abord, pas de validation de type.
        """
        langs = ("fr", "en")
        validate = entity_type in ("ORG", "PRODUCT")

        for lang in langs:
            try:
                r = req.get(
                    f"https://{lang}.wikipedia.org/w/api.php",
                    params={
                        "action": "query",
                        "generator": "search",
                        "gsrsearch": name,
                        "gsrlimit": 3,
                        "prop": "pageimages|pageprops",
                        "pithumbsize": THUMB,
                        "pilicense": "any",
                        "ppprop": "wikibase_item",
                        "format": "json",
                        "origin": "*",
                    },
                    headers={"User-Agent": UA},
                    timeout=8,
                )
                pages = r.json().get("query", {}).get("pages", {})
                for page in sorted(pages.values(), key=lambda p: p.get("index", 0)):
                    # Validation du type Wikidata pour ORG/PRODUCT
                    if validate:
                        qid = page.get("pageprops", {}).get("wikibase_item", "")
                        if not _wikidata_type_ok(qid, strict=True):
                            continue  # mauvais type ou P31 absent → essai page suivante

                    thumb = page.get("thumbnail")
                    if thumb and thumb.get("source"):
                        return name, {
                            "url": thumb["source"],
                            "width": thumb.get("width", THUMB),
                            "height": thumb.get("height", THUMB),
                        }
            except Exception:
                continue
        return name, None

    # ── PERSON & autres : pageimages puis Wikidata P18 si rien trouvé ───────────
    pageimg = _pageimages(person_names + other_names)
    for name in person_names + other_names:
        if name not in cache:
            cache[name] = pageimg.get(name)

    # Fallback Wikidata P18 pour les PERSON sans image Wikipedia
    persons_no_img = [n for n in person_names if not cache.get(n)]
    if persons_no_img:
        p18_files = _wikidata_p18_persons(persons_no_img)
        if p18_files:
            p18_urls = _resolve_logo_urls(list(set(p18_files.values())))
            for name, fname in p18_files.items():
                if name not in cache or not cache[name]:
                    url = p18_urls.get(fname)
                    cache[name] = {"url": url, "width": THUMB, "height": THUMB} if url else None

    # ── ORG / PRODUCT : Wikidata P154/P18 → sinon _search_pageimage_single ──
    if logo_names:
        wikidata, rejected = _wikidata_logos(logo_names)
        # Résout les URLs des logos trouvés via P154
        resolved = _resolve_logo_urls(list(set(wikidata.values()))) if wikidata else {}
        for name in logo_names:
            if name not in cache:
                logo_file = wikidata.get(name)
                if logo_file and logo_file in resolved:
                    cache[name] = {"url": resolved[logo_file], "width": THUMB, "height": THUMB}
                elif name in rejected:
                    cache[name] = None  # type hors-scope confirmé (personne, prénom, homonymie) → pas d'image
                # else: pas de logo Wikidata → sera traité par _search_pageimage_single

    # ── Fallback final : Wikipedia generator=search pour toutes les entités sans image ──
    SEARCH_LIMIT = 25  # max entités par passe pour limiter la latence
    _rejected = rejected if logo_names else set()
    _type_map = {e["name"]: e["type"] for e in to_fetch}
    null_entities = [
        name for name in (person_names + logo_names + other_names)
        if not cache.get(name) and name not in _rejected
    ][:SEARCH_LIMIT]
    if null_entities:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = {
                pool.submit(_search_pageimage_single, name, _type_map.get(name, "")): name
                for name in null_entities
            }
            for future in as_completed(futures):
                try:
                    name, result = future.result()
                    if result and not cache.get(name):
                        cache[name] = result
                except Exception:
                    pass

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")

    return jsonify({e["name"]: cache.get(e["name"]) for e in entities})




@app.route("/api/entities/info")
def api_entities_info():
    """Génère en streaming une synthèse encyclopédique sur une entité (EurIA ou Claude)."""
    import requests as req

    entity_type  = request.args.get("type",  "").strip()
    entity_value = request.args.get("value", "").strip()
    if not entity_type or not entity_value:
        return jsonify({"error": "Paramètres type et value requis"}), 400

    provider = os.environ.get("AI_PROVIDER", "euria").strip().lower()

    type_labels = {
        "PERSON":      "personne physique",
        "ORG":         "organisation ou entreprise",
        "GPE":         "lieu géopolitique",
        "LOC":         "lieu géographique",
        "PRODUCT":     "produit ou technologie",
        "EVENT":       "événement",
        "WORK_OF_ART": "œuvre",
        "LAW":         "loi ou règlement",
        "NORP":        "groupe national, religieux ou politique",
        "FAC":         "site ou bâtiment",
    }
    label = type_labels.get(entity_type, entity_type.lower())

    prompt = (
        f"Fournis une synthèse encyclopédique en français sur « {entity_value} » ({label}).\n\n"
        "Structure ta réponse en Markdown avec des sections pertinentes "
        "(présentation, rôle, contexte, actualité récente, chiffres clés, liens avec d'autres acteurs…).\n"
        "Sois factuel et concis. Génère uniquement le contenu Markdown, sans balises <think>."
    )

    if provider == "claude":
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"error": "ANTHROPIC_API_KEY manquante dans .env (AI_PROVIDER=claude)"}), 503
        from utils.api_client import ClaudeClient as _ClaudeClient
        _claude = _ClaudeClient(api_key=api_key)

        def generate():
            yield from _claude.stream(prompt=prompt, timeout=90)

    else:
        api_url = os.environ.get("URL", "")
        bearer  = os.environ.get("bearer", "")
        if not api_url or not bearer:
            return jsonify({"error": "URL ou bearer manquant dans .env (AI_PROVIDER=euria)"}), 503
        payload = {
            "messages": [{"role": "user", "content": prompt}],
            "model": "qwen3",
            "stream": True,
            "enable_web_search": True,
        }
        api_headers = {
            "Authorization": f"Bearer {bearer}",
            "Content-Type": "application/json",
        }

        def generate():
            try:
                r = req.post(api_url, json=payload, headers=api_headers, stream=True, timeout=90)
                r.raise_for_status()
                for line in r.iter_lines():
                    if line:
                        yield line.decode("utf-8") + "\n\n"
            except Exception as exc:
                yield f'data: {json.dumps({"error": str(exc)})}\n\n'

    return Response(
        stream_with_context(generate()),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Scripts ───────────────────────────────────────────────────────────────────

@app.route("/api/scripts/keyword-rss/status")
def keyword_rss_status():
    """Retourne l'état courant du script get-keyword-from-rss.py.

    Combine trois sources :
    1. _rss_job   — suivi du process lancé via le viewer
    2. ps aux     — détection d'un process externe (cron)
    3. rss_progress.json — fichier de progression écrit par le script lui-même
    """
    # ── 1. Process suivi par le viewer ────────────────────────────────────────
    with _rss_job["lock"]:
        proc = _rss_job.get("process")
        viewer_running = proc is not None and proc.poll() is None
        viewer_pid     = proc.pid if viewer_running and proc else None
        last_run       = _rss_job.get("last_run")
        last_rc        = _rss_job.get("last_returncode")

    # ── 2. Détection d'un process externe (cron / terminal) ──────────────────
    external_pid = None
    try:
        out = subprocess.check_output(
            ["pgrep", "-f", "get-keyword-from-rss"],
            text=True, stderr=subprocess.DEVNULL
        ).strip()
        pids = [int(p) for p in out.splitlines() if p.strip().isdigit()]
        # Exclure le PID du viewer lui-même et de son process enfant (si lancé via viewer)
        for p in pids:
            if viewer_pid and p == viewer_pid:
                continue
            external_pid = p
            break
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass

    running = viewer_running or (external_pid is not None)
    pid     = viewer_pid or external_pid

    # ── 3. Fichier de progression écrit par le script ─────────────────────────
    progress = None
    progress_file = PROJECT_ROOT / "data" / "rss_progress.json"
    if progress_file.exists():
        try:
            progress = json.loads(progress_file.read_text(encoding="utf-8"))
        except Exception:
            pass

    # Réconcilier avec rss_progress.json :
    # - Si started_at présent, finished_at absent, et fichier modifié < 3 min → script en cours
    # - Si finished_at présent → terminé, récupère last_run et returncode
    if progress:
        if progress.get("started_at") and not progress.get("finished_at") and not running:
            try:
                mtime = progress_file.stat().st_mtime
                age_s = datetime.datetime.now().timestamp() - mtime
                if age_s < 180:  # fichier modifié il y a moins de 3 minutes
                    running = True
            except Exception:
                pass
        elif progress.get("finished_at") and not running:
            last_run = last_run or progress.get("finished_at")
            last_rc  = last_rc if last_rc is not None else progress.get("returncode")

    # ── 4. Comptage des fichiers et articles dans articles-from-rss ──────────
    articles_dir = PROJECT_ROOT / "data" / "articles-from-rss"
    file_count    = 0
    article_count = 0
    if articles_dir.exists():
        for f in articles_dir.glob("*.json"):
            file_count += 1
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    article_count += len(data)
            except Exception:
                pass

    result = {
        "running":          running,
        "pid":              pid,
        "last_run":         last_run,
        "last_returncode":  last_rc,
        "file_count":       file_count,
        "article_count":    article_count,
        "progress":         progress,  # None si le script n'a jamais écrit le fichier
    }
    return jsonify(result)


@app.route("/api/scripts/keyword-rss/stream")
def stream_keyword_rss():
    """Lance get-keyword-from-rss.py et stream la sortie via SSE."""
    script = PROJECT_ROOT / "scripts" / "get-keyword-from-rss.py"
    if not script.exists():
        return jsonify({"error": f"Script introuvable : {script}"}), 404

    def generate():
        with _rss_job["lock"]:
            proc = _rss_job["process"]
            if proc is not None and proc.poll() is None:
                yield f'data: {json.dumps({"log": f"⚠ Script déjà en cours (PID {proc.pid})"})}\n\n'
                return
            env = {**os.environ, "PYTHONUNBUFFERED": "1"}
            try:
                proc = subprocess.Popen(
                    ["python3", str(script)],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                    cwd=str(PROJECT_ROOT),
                    env=env,
                )
                _rss_job["process"] = proc
            except Exception as exc:
                yield f'data: {json.dumps({"error": str(exc)})}\n\n'
                return

        yield f'data: {json.dumps({"log": f"▶ Démarré (PID {proc.pid})"})}\n\n'

        try:
            for line in proc.stdout:
                stripped = line.rstrip("\n")
                if stripped:
                    yield f'data: {json.dumps({"log": stripped})}\n\n'
        except Exception as exc:
            yield f'data: {json.dumps({"error": str(exc)})}\n\n'

        rc = proc.wait()
        with _rss_job["lock"]:
            _rss_job["last_run"] = datetime.datetime.now(datetime.timezone.utc).isoformat()
            _rss_job["last_returncode"] = rc
        msg = f"✓ Terminé (code : {rc})" if rc == 0 else f"✗ Erreur (code : {rc})"
        yield f'data: {json.dumps({"done": True, "returncode": rc, "log": msg})}\n\n'

    return Response(
        stream_with_context(generate()),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/files", methods=["DELETE"])
def api_delete_file():
    """Supprime un fichier de data/ ou rapports/ (avec validation de sécurité)."""
    rel = request.args.get("path", "").strip()
    if not rel:
        abort(400, "Paramètre path requis")
    if not (rel.startswith("data/") or rel.startswith("rapports/")):
        abort(403, "Suppression non autorisée hors de data/ et rapports/")
    target = (PROJECT_ROOT / rel).resolve()
    if not str(target).startswith(str(PROJECT_ROOT) + "/"):
        abort(403, "Accès refusé")
    if not target.exists():
        abort(404, "Fichier non trouvé")
    try:
        target.unlink()
    except OSError as exc:
        abort(500, f"Erreur de suppression : {exc}")
    return jsonify({"ok": True, "deleted": rel})


# ── Serveur frontend React (production) ──────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════
# NOUVELLES FEATURES (v2.2) — Scoring, Tendances, Sentiment, RAG, Export
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/api/articles/top")
def api_articles_top():
    """Retourne les N articles les mieux scorés sur une fenêtre temporelle.

    Paramètres :
      n     : nombre d'articles (défaut: 10, max: 50)
      hours : fenêtre en heures (défaut: 48, 0=sans filtre)
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT))
    try:
        from utils.scoring import ScoringEngine
        n = min(int(request.args.get("n", 10)), 50)
        hours = int(request.args.get("hours", 48))
        engine = ScoringEngine(PROJECT_ROOT)
        top = engine.get_top_articles(top_n=n, hours=hours)
        return jsonify(top)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/alerts")
def api_get_alerts():
    """Retourne les alertes de tendance (data/alertes.json).

    Paramètres :
      niveau : filtre par niveau ("critique", "élevé", "modéré")
    """
    alerts_file = PROJECT_ROOT / "data" / "alertes.json"
    if not alerts_file.exists():
        return jsonify([])
    try:
        alerts = json.loads(alerts_file.read_text(encoding="utf-8"))
        niveau = request.args.get("niveau", "").strip()
        if niveau:
            alerts = [a for a in alerts if a.get("niveau") == niveau]
        return jsonify(alerts)
    except (json.JSONDecodeError, OSError) as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/alerts/run", methods=["POST"])
def api_run_trend_detector():
    """Lance le détecteur de tendances et retourne les alertes générées."""
    import sys
    import subprocess
    script = PROJECT_ROOT / "scripts" / "trend_detector.py"
    if not script.exists():
        return jsonify({"error": "Script trend_detector.py introuvable"}), 404

    data = request.get_json(force=True) or {}
    threshold = float(data.get("threshold", 2.0))
    top = int(data.get("top", 20))

    try:
        result = subprocess.run(
            [sys.executable, str(script), "--threshold", str(threshold), "--top", str(top)],
            capture_output=True, text=True, timeout=120, cwd=str(PROJECT_ROOT)
        )
        alerts_file = PROJECT_ROOT / "data" / "alertes.json"
        alerts = []
        if alerts_file.exists():
            try:
                alerts = json.loads(alerts_file.read_text(encoding="utf-8"))
            except Exception:
                pass
        return jsonify({
            "ok": result.returncode == 0,
            "alerts": alerts,
            "stdout": result.stdout[-2000:] if result.stdout else "",
        })
    except subprocess.TimeoutExpired:
        return jsonify({"error": "Timeout (120s)"}), 504
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/sources/bias")
def api_sources_bias():
    """Agrège les données de sentiment par source pour détecter les biais éditoriaux.

    Returns :
      [{ source, article_count, sentiment_counts: {positif, neutre, négatif},
         avg_score_sentiment, avg_score_ton, ton_distribution: {...} }]
    """
    from collections import defaultdict

    data_dirs = [
        PROJECT_ROOT / "data" / "articles",
        PROJECT_ROOT / "data" / "articles-from-rss",
    ]

    sources: dict[str, dict] = defaultdict(lambda: {
        "article_count": 0,
        "sentiment_counts": {"positif": 0, "neutre": 0, "négatif": 0},
        "score_sentiment_sum": 0,
        "score_ton_sum": 0,
        "score_count": 0,
        "ton_distribution": {},
    })

    for data_dir in data_dirs:
        if not data_dir.exists():
            continue
        for json_file in data_dir.rglob("*.json"):
            if "cache" in str(json_file):
                continue
            try:
                articles = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if not isinstance(articles, list):
                    continue
            except (json.JSONDecodeError, OSError):
                continue
            for article in articles:
                source = article.get("Sources", "Inconnu").strip()
                if not source:
                    continue
                s = sources[source]
                s["article_count"] += 1
                sentiment = article.get("sentiment", "")
                if sentiment in ("positif", "neutre", "négatif"):
                    s["sentiment_counts"][sentiment] += 1
                score_s = article.get("score_sentiment")
                score_t = article.get("score_ton")
                if isinstance(score_s, (int, float)) and isinstance(score_t, (int, float)):
                    s["score_sentiment_sum"] += score_s
                    s["score_ton_sum"] += score_t
                    s["score_count"] += 1
                ton = article.get("ton_editorial", "")
                if ton:
                    s["ton_distribution"][ton] = s["ton_distribution"].get(ton, 0) + 1

    result = []
    for source, data in sources.items():
        count = data["score_count"]
        result.append({
            "source": source,
            "article_count": data["article_count"],
            "sentiment_counts": data["sentiment_counts"],
            "avg_score_sentiment": round(data["score_sentiment_sum"] / count, 2) if count else None,
            "avg_score_ton": round(data["score_ton_sum"] / count, 2) if count else None,
            "ton_distribution": data["ton_distribution"],
        })

    result.sort(key=lambda x: x["article_count"], reverse=True)
    return jsonify(result)


@app.route("/api/synthesize-topic")
def api_synthesize_topic():
    """Synthèse comparative multi-sources en streaming SSE.

    Paramètres :
      entity_type  : type de l'entité (ex: "ORG")
      entity_value : valeur de l'entité (ex: "OpenAI")
      topic        : sujet libre (alternatif à entity_type+entity_value)
      n            : nombre d'articles à consolider (défaut: 15)
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT))

    entity_type = request.args.get("entity_type", "").strip()
    entity_value = request.args.get("entity_value", "").strip()
    topic = request.args.get("topic", "").strip()
    n = min(int(request.args.get("n", 15)), 30)

    if not topic and not entity_value:
        return jsonify({"error": "Paramètre topic ou entity_value requis"}), 400

    label = topic or entity_value

    # Collecte des articles pertinents
    matching_articles = []
    search_term = (entity_value or topic).lower()

    for data_dir in [PROJECT_ROOT / "data" / "articles", PROJECT_ROOT / "data" / "articles-from-rss"]:
        if not data_dir.exists():
            continue
        for json_file in sorted(data_dir.rglob("*.json")):
            if "cache" in str(json_file):
                continue
            try:
                arts = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if not isinstance(arts, list):
                    continue
            except (json.JSONDecodeError, OSError):
                continue
            for article in arts:
                resume = (article.get("Résumé") or "").lower()
                entities = article.get("entities", {})
                # Correspondance : entité NER OU présence dans le résumé
                entity_match = False
                if entity_type and entity_value:
                    values = entities.get(entity_type, []) if isinstance(entities, dict) else []
                    entity_match = entity_value in values
                text_match = search_term in resume
                if entity_match or text_match:
                    matching_articles.append(article)

    # Déduplication par URL
    seen_urls = set()
    deduped = []
    for a in matching_articles:
        url = a.get("URL", "")
        if url and url in seen_urls:
            continue
        if url:
            seen_urls.add(url)
        deduped.append(a)

    # Tri par date décroissante
    deduped.sort(key=lambda a: a.get("Date de publication", ""), reverse=True)
    articles_to_use = deduped[:n]

    provider = os.environ.get("AI_PROVIDER", "euria").strip().lower()

    if provider == "claude":
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"error": "ANTHROPIC_API_KEY manquante dans .env (AI_PROVIDER=claude)"}), 503
    else:
        api_url = os.environ.get("URL", "")
        bearer  = os.environ.get("bearer", "")
        if not api_url or not bearer:
            return jsonify({"error": "URL ou bearer manquant dans .env (AI_PROVIDER=euria)"}), 503

    if not articles_to_use:
        def empty_stream():
            msg = f"Aucun article trouvé pour « {label} »."
            yield f'data: {json.dumps({"choices":[{"delta":{"content": msg},"finish_reason":None}]})}\n\n'
            yield "data: [DONE]\n\n"
        return Response(stream_with_context(empty_stream()), content_type="text/event-stream",
                        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    # Construire le prompt à partir des articles collectés
    sources_block = ""
    for i, a in enumerate(articles_to_use, 1):
        source = a.get("Sources", "Source inconnue")
        date   = a.get("Date de publication", "")
        resume = (a.get("Résumé") or "")[:600]
        sources_block += f"\n--- Article {i} ({source}, {date}) ---\n{resume}\n"

    prompt = (
        f"Tu es un analyste de presse. Voici {len(articles_to_use)} articles de sources différentes "
        f"traitant du sujet : **{label}**.\n\n"
        "Génère une synthèse comparative structurée en Markdown comprenant :\n"
        "1. **Résumé de la situation** (2-3 phrases)\n"
        "2. **Points de convergence** entre les sources\n"
        "3. **Points de divergence ou contradictions**\n"
        "4. **Positionnement éditorial** : sources favorables, neutres ou critiques\n"
        "5. **Éléments clés manquants**\n\n"
        "Cite les sources (nom + date) à chaque point. Sois concis et factuel.\n"
        "Génère uniquement le contenu Markdown, sans balises <think>.\n\n"
        f"Articles :\n{sources_block}"
    )

    import requests as req

    if provider == "claude":
        from utils.api_client import ClaudeClient as _ClaudeClient
        _claude = _ClaudeClient(api_key=api_key)

        def generate_synthesis():
            yield from _claude.stream(prompt=prompt, timeout=120)

    else:
        payload = {
            "messages": [{"role": "user", "content": prompt}],
            "model": "qwen3",
            "stream": True,
        }
        api_headers = {
            "Authorization": f"Bearer {bearer}",
            "Content-Type": "application/json",
        }

        def generate_synthesis():
            try:
                r = req.post(api_url, json=payload, headers=api_headers, stream=True, timeout=120)
                r.raise_for_status()
                for line in r.iter_lines():
                    if line:
                        decoded = line.decode("utf-8")
                        # Normalise le préfixe SSE : certaines versions de l'API
                        # envoient du JSON brut sans "data: "
                        if not decoded.startswith("data:"):
                            decoded = "data: " + decoded
                        yield decoded + "\n\n"
            except Exception as exc:
                yield f'data: {json.dumps({"error": str(exc)})}\n\n'

    return Response(
        stream_with_context(generate_synthesis()),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/export/atom")
def api_export_atom():
    """Génère et retourne un flux Atom pour un flux ou tous les articles.

    Paramètres :
      flux        : nom du flux (ex: "Intelligence-artificielle")
      keyword     : mot-clé (ex: "OpenAI")
      max_entries : nombre max d'entrées (défaut: 50)
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT))
    try:
        from utils.exporters.atom_feed import generate_atom_feed, generate_atom_from_flux

        flux = request.args.get("flux", "").strip()
        keyword = request.args.get("keyword", "").strip()
        max_entries = min(int(request.args.get("max_entries", 50)), 200)

        # URL canonique dynamique : utilise l'URL réelle de la requête (évite localhost vs IP réseau)
        actual_self_url = request.url

        if flux:
            xml = generate_atom_from_flux(PROJECT_ROOT, flux, max_entries, self_url=actual_self_url)
            feed_title = f"WUDD.ai · {flux}"
        elif keyword:
            kw_file = PROJECT_ROOT / "data" / "articles-from-rss" / f"{keyword}.json"
            if not kw_file.exists():
                return jsonify({"error": "Fichier keyword introuvable"}), 404
            articles = json.loads(kw_file.read_text(encoding="utf-8"))
            articles.sort(key=lambda a: a.get("Date de publication", ""), reverse=True)
            from utils.exporters.atom_feed import _FEED_ID_BASE
            xml = generate_atom_feed(
                articles, feed_title=f"WUDD.ai · {keyword}",
                feed_id=f"{_FEED_ID_BASE}keyword-{keyword.lower()}",
                self_url=actual_self_url,
                max_entries=max_entries,
            )
        else:
            # Tout agréger
            all_articles = []
            for d in [PROJECT_ROOT / "data" / "articles", PROJECT_ROOT / "data" / "articles-from-rss"]:
                if not d.exists():
                    continue
                for jf in sorted(d.rglob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True)[:10]:
                    if "cache" in str(jf):
                        continue
                    try:
                        data = json.loads(jf.read_text(encoding="utf-8"))
                        if isinstance(data, list):
                            all_articles.extend(data)
                    except Exception:
                        continue
            all_articles.sort(key=lambda a: a.get("Date de publication", ""), reverse=True)
            xml = generate_atom_feed(all_articles, feed_title="WUDD.ai · Veille complète",
                                     self_url=actual_self_url,
                                     max_entries=max_entries)

        return Response(xml, mimetype="application/atom+xml; charset=utf-8")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/newsletter", methods=["GET", "POST"])
def api_export_newsletter():
    """Génère une newsletter HTML depuis les articles récents.

    GET  → retourne le HTML brut
    POST → { send: true } pour envoyer par SMTP (si configuré)

    Paramètres GET :
      hours : fenêtre temporelle (défaut: 48)
      title : titre de la newsletter
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT))
    try:
        from utils.exporters.newsletter import generate_newsletter_html, send_newsletter
        from utils.scoring import ScoringEngine

        hours = int(request.args.get("hours", 48))
        title = request.args.get("title", "").strip() or \
            f"Veille WUDD.ai — {datetime.datetime.now().strftime('%d %B %Y')}"

        engine = ScoringEngine(PROJECT_ROOT)
        articles = engine.get_top_articles(top_n=20, hours=hours)
        html = generate_newsletter_html(articles, title=title)

        # Sauvegarde locale
        nl_dir = PROJECT_ROOT / "rapports" / "html"
        nl_dir.mkdir(parents=True, exist_ok=True)
        slug = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
        nl_file = nl_dir / f"newsletter_{slug}.html"
        nl_file.write_text(html, encoding="utf-8")

        if request.method == "POST":
            data = request.get_json(force=True) or {}
            if data.get("send"):
                success = send_newsletter(html, subject=title)
                return jsonify({"ok": success, "path": str(nl_file.relative_to(PROJECT_ROOT))})

        return Response(html, mimetype="text/html; charset=utf-8")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/webhook-test", methods=["POST"])
def api_webhook_test():
    """Teste l'envoi webhook avec les alertes actuelles.

    Body JSON : { platform: "discord"|"slack"|"ntfy"|"all" }
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT))
    try:
        from utils.exporters.webhook import send_discord, send_slack, send_ntfy, notify_alerts

        platform = (request.get_json(force=True) or {}).get("platform", "all")
        alerts_file = PROJECT_ROOT / "data" / "alertes.json"
        alerts = []
        if alerts_file.exists():
            try:
                alerts = json.loads(alerts_file.read_text(encoding="utf-8"))
            except Exception:
                pass

        if not alerts:
            return jsonify({"ok": False, "message": "Aucune alerte disponible — lancez d'abord trend_detector.py"})

        if platform == "discord":
            ok = send_discord(alerts)
            return jsonify({"discord": ok})
        elif platform == "slack":
            ok = send_slack(alerts)
            return jsonify({"slack": ok})
        elif platform == "ntfy":
            ok = send_ntfy(alerts)
            return jsonify({"ntfy": ok})
        else:
            results = notify_alerts(alerts)
            return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Nouvelles fonctionnalités (v2.3) ─────────────────────────────────────────

@app.route("/api/entities/timeline")
def api_entities_timeline():
    """Série chronologique des mentions d'entités.

    Query params :
      days       : fenêtre temporelle en jours (défaut 30)
      top        : nombre d'entités (défaut 30)
      entity     : filtrer sur une valeur d'entité
      type       : filtrer sur un type d'entité (PERSON, ORG…)
      regenerate : si "1", force le recalcul (sinon utilise le cache JSON)
    """
    try:
        days       = int(request.args.get("days", 30))
        top_n      = int(request.args.get("top", 30))
        entity     = request.args.get("entity") or None
        etype      = request.args.get("type")   or None
        regenerate = request.args.get("regenerate") == "1"

        timeline_file = PROJECT_ROOT / "data" / "entity_timeline.json"

        # Utiliser le fichier mis en cache si présent et non périmé (< 1h)
        if not regenerate and timeline_file.exists() and not entity and not etype:
            import time as _time
            age_s = _time.time() - timeline_file.stat().st_mtime
            if age_s < 3600:
                data = json.loads(timeline_file.read_text(encoding="utf-8"))
                return jsonify(data)

        # Import local pour éviter la dépendance circulaire au démarrage
        import sys as _sys
        _sys.path.insert(0, str(PROJECT_ROOT))
        from scripts.entity_timeline import collect_timeline, fill_missing_dates, build_top_entities

        raw = collect_timeline(PROJECT_ROOT, days=days, entity_filter=entity, type_filter=etype)
        top_entities = build_top_entities(raw, top_n=top_n)
        top_keys = {e["key"] for e in top_entities}
        filled = fill_missing_dates({k: v for k, v in raw.items() if k in top_keys}, days=days)

        result = {
            "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
            "window_days": days,
            "top_entities": top_entities,
            "timeline": filled,
        }

        # Sauvegarder le cache si requête sans filtre
        if not entity and not etype:
            timeline_file.parent.mkdir(parents=True, exist_ok=True)
            timeline_file.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

        return jsonify(result)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/sources/credibility")
def api_sources_credibility():
    """Score de crédibilité des sources.

    Query params :
      source : nom de la source à évaluer (retourne une seule entrée)
               Si absent, retourne toutes les sources de la base
    """
    try:
        from utils.source_credibility import CredibilityEngine
        engine = CredibilityEngine(PROJECT_ROOT)

        source_query = request.args.get("source") or None
        if source_query:
            meta = engine.get_metadata(source_query)
            meta["source"] = source_query
            meta["multiplier"] = engine.get_multiplier(source_query)
            return jsonify(meta)

        # Toutes les sources de la base
        all_sources = []
        for name, entry in engine._db.items():
            all_sources.append({
                "source":     name,
                "score":      entry.get("score", 50),
                "biais":      entry.get("biais", "inconnu"),
                "type":       entry.get("type", "inconnu"),
                "pays":       entry.get("pays", "inconnu"),
                "fiabilite":  entry.get("fiabilite", "non évalué"),
                "multiplier": engine.get_multiplier(name),
            })
        all_sources.sort(key=lambda x: -x["score"])
        return jsonify({"sources": all_sources, "total": len(all_sources)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/alerts/rules", methods=["GET"])
def api_get_alert_rules():
    """Retourne la configuration des règles d'alertes."""
    rules_file = PROJECT_ROOT / "config" / "alert_rules.json"
    if not rules_file.exists():
        return jsonify({"error": "alert_rules.json introuvable"}), 404
    try:
        return jsonify(json.loads(rules_file.read_text(encoding="utf-8")))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/alerts/rules", methods=["POST"])
def api_save_alert_rules():
    """Sauvegarde la configuration des règles d'alertes."""
    rules_file = PROJECT_ROOT / "config" / "alert_rules.json"
    try:
        data = request.get_json(force=True)
        if not isinstance(data, dict):
            return jsonify({"error": "Données invalides (dict attendu)"}), 400
        rules_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return jsonify({"status": "ok", "message": "Règles d'alertes sauvegardées"})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/briefing/generate", methods=["POST"])
def api_generate_briefing():
    """Génère un briefing exécutif (sans synthèse IA, retourne le Markdown).

    Body JSON :
      period : "daily" (défaut) ou "weekly"
    """
    try:
        body    = request.get_json(force=True) or {}
        period  = body.get("period", "daily")
        if period not in ("daily", "weekly"):
            return jsonify({"error": "period doit être 'daily' ou 'weekly'"}), 400

        import sys as _sys
        _sys.path.insert(0, str(PROJECT_ROOT))
        from scripts.generate_briefing import (
            collect_articles, compute_top_entities, load_alerts,
            build_briefing_markdown, _PERIOD_HOURS,
        )
        from utils.scoring import ScoringEngine
        from datetime import timedelta

        hours = _PERIOD_HOURS[period]
        now   = datetime.datetime.utcnow()
        date_fin   = now.strftime("%Y-%m-%d")
        date_debut = (now - timedelta(hours=hours)).strftime("%Y-%m-%d")
        period_label = "hebdomadaire" if period == "weekly" else "quotidien"

        articles     = collect_articles(PROJECT_ROOT, hours=hours)
        engine       = ScoringEngine(PROJECT_ROOT)
        top_articles = engine.score_and_sort(articles, top_n=10)
        top_entities = compute_top_entities(articles, top_n=10)
        alerts       = load_alerts(PROJECT_ROOT)

        md = build_briefing_markdown(
            period_label=period_label,
            date_debut=date_debut,
            date_fin=date_fin,
            articles=articles,
            top_articles=top_articles,
            top_entities=top_entities,
            alerts=alerts,
        )
        return jsonify({
            "period":       period,
            "date_debut":   date_debut,
            "date_fin":     date_fin,
            "articles_count": len(articles),
            "markdown":     md,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/cross-flux")
def api_cross_flux():
    """Analyse croisée des flux — entités communes.

    Query params :
      days     : fenêtre temporelle en jours (défaut 30)
      min_flux : nombre minimal de flux (défaut 2)
      top      : nombre d'entités (défaut 30)
    """
    try:
        days      = int(request.args.get("days", 30))
        min_flux  = int(request.args.get("min_flux", 2))
        top_n     = int(request.args.get("top", 30))

        # Essayer d'abord le fichier mis en cache
        cache_file = PROJECT_ROOT / "data" / "cross_flux_report.json"
        if cache_file.exists():
            import time as _time
            age_s = _time.time() - cache_file.stat().st_mtime
            if age_s < 3600:
                data = json.loads(cache_file.read_text(encoding="utf-8"))
                return jsonify(data)

        import sys as _sys
        _sys.path.insert(0, str(PROJECT_ROOT))
        from scripts.cross_flux_analysis import collect_entities_by_flux, compute_cross_flux

        flux_entities   = collect_entities_by_flux(PROJECT_ROOT, days=days)
        cross_entities  = compute_cross_flux(flux_entities, min_flux=min_flux, top_n=top_n)

        result = {
            "generated_at":   datetime.datetime.utcnow().isoformat() + "Z",
            "window_days":    days,
            "min_flux":       min_flux,
            "flux_count":     len(flux_entities),
            "flux_list":      sorted(flux_entities.keys()),
            "cross_entities": cross_entities,
        }

        # Mettre en cache
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        cache_file.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

        return jsonify(result)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ── Annotations manuelles ─────────────────────────────────────────────────────
# Stockées dans data/annotations.json (dict keyed par URL d'article)
# Jamais dans les fichiers articles — données sources préservées.

_ANNOTATIONS_FILE = PROJECT_ROOT / "data" / "annotations.json"
_annotations_lock = threading.Lock()

def _load_annotations() -> dict:
    """Charge le fichier annotations.json (crée s'il n'existe pas)."""
    if not _ANNOTATIONS_FILE.exists():
        return {}
    try:
        return json.loads(_ANNOTATIONS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _save_annotations(data: dict) -> None:
    """Sauvegarde atomique du fichier annotations.json."""
    _ANNOTATIONS_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = _ANNOTATIONS_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(_ANNOTATIONS_FILE)


@app.route("/api/annotations", methods=["GET"])
def api_annotations_get():
    """Retourne toutes les annotations (dict keyed par URL)."""
    with _annotations_lock:
        return jsonify(_load_annotations())


@app.route("/api/annotations", methods=["POST"])
def api_annotations_post():
    """Crée ou met à jour l'annotation d'un article.

    Body JSON attendu :
        url         (str, obligatoire) — URL de l'article
        is_important (bool, optionnel)
        is_read      (bool, optionnel)
        tags         (list[str], optionnel, max 20 items)
        notes        (str, optionnel, max 5000 chars)
    """
    body = request.get_json(force=True, silent=True) or {}
    url = (body.get("url") or "").strip()
    if not url:
        return jsonify({"error": "Le champ 'url' est obligatoire"}), 400

    now_iso = datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")

    with _annotations_lock:
        data = _load_annotations()
        existing = data.get(url, {})

        # Merge : on ne remplace que les champs explicitement fournis
        updated = dict(existing)
        if "is_important" in body:
            updated["is_important"] = bool(body["is_important"])
        if "is_read" in body:
            updated["is_read"] = bool(body["is_read"])
        if "tags" in body:
            tags = body["tags"]
            if not isinstance(tags, list):
                return jsonify({"error": "'tags' doit être une liste"}), 400
            tags = [str(t).strip() for t in tags if str(t).strip()][:20]
            updated["tags"] = tags
        if "notes" in body:
            notes = str(body["notes"])[:5000]
            updated["notes"] = notes

        updated["updated_at"] = now_iso
        if "created_at" not in updated:
            updated["created_at"] = now_iso

        data[url] = updated
        _save_annotations(data)

    return jsonify({"ok": True, "url": url, "annotation": updated})


@app.route("/api/annotations", methods=["DELETE"])
def api_annotations_delete():
    """Supprime l'annotation d'un article (paramètre ?url=...)."""
    url = (request.args.get("url") or "").strip()
    if not url:
        return jsonify({"error": "Paramètre 'url' obligatoire"}), 400

    with _annotations_lock:
        data = _load_annotations()
        if url not in data:
            return jsonify({"ok": True, "removed": False})
        del data[url]
        _save_annotations(data)

    return jsonify({"ok": True, "removed": True, "url": url})


# ── Entités surveillées ───────────────────────────────────────────────────────
# Stockées dans data/watched_entities.json
# [{type, value, added_at, notes}]

_WATCHED_FILE = PROJECT_ROOT / "data" / "watched_entities.json"
_watched_lock = threading.Lock()


def _load_watched() -> list:
    if not _WATCHED_FILE.exists():
        return []
    try:
        return json.loads(_WATCHED_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def _save_watched(data: list) -> None:
    _WATCHED_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = _WATCHED_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(_WATCHED_FILE)


@app.route("/api/watched-entities", methods=["GET"])
def api_watched_get():
    """Retourne les entités surveillées avec leur volume de mentions récentes."""
    with _watched_lock:
        watched = _load_watched()

    # Calcul rapide des mentions sur les 7 derniers jours
    from datetime import datetime, timedelta, timezone
    now = datetime.now(timezone.utc)
    cutoff_7d = now - timedelta(days=7)
    cutoff_24h = now - timedelta(hours=24)

    counts_7d: dict[str, int] = {}
    counts_24h: dict[str, int] = {}

    for data_dir in [PROJECT_ROOT / "data" / "articles", PROJECT_ROOT / "data" / "articles-from-rss"]:
        if not data_dir.exists():
            continue
        for json_file in data_dir.rglob("*.json"):
            if "cache" in str(json_file):
                continue
            try:
                arts = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if not isinstance(arts, list):
                    continue
            except (json.JSONDecodeError, OSError):
                continue
            for art in arts:
                entities = art.get("entities", {})
                if not isinstance(entities, dict):
                    continue
                # Parse date
                date_str = art.get("Date de publication", "")
                art_dt = None
                for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d", "%d/%m/%Y"):
                    try:
                        art_dt = datetime.strptime(date_str[:19], fmt).replace(tzinfo=timezone.utc)
                        break
                    except ValueError:
                        continue
                if art_dt is None:
                    try:
                        from email.utils import parsedate_to_datetime
                        art_dt = parsedate_to_datetime(date_str).astimezone(timezone.utc)
                    except Exception:
                        pass

                for w in watched:
                    vals = entities.get(w["type"], [])
                    if isinstance(vals, list) and w["value"] in vals:
                        key = f"{w['type']}:{w['value']}"
                        if art_dt and art_dt >= cutoff_7d:
                            counts_7d[key] = counts_7d.get(key, 0) + 1
                        if art_dt and art_dt >= cutoff_24h:
                            counts_24h[key] = counts_24h.get(key, 0) + 1

    result = []
    for w in watched:
        key = f"{w['type']}:{w['value']}"
        result.append({**w, "mentions_7d": counts_7d.get(key, 0), "mentions_24h": counts_24h.get(key, 0)})

    return jsonify(result)


@app.route("/api/watched-entities", methods=["POST"])
def api_watched_post():
    """Ajoute ou met à jour une entité surveillée.

    Body JSON : { type: str, value: str, notes?: str }
    """
    body = request.get_json(force=True, silent=True) or {}
    etype = (body.get("type") or "").strip().upper()
    value = (body.get("value") or "").strip()
    if not etype or not value:
        return jsonify({"error": "Champs type et value requis"}), 400

    with _watched_lock:
        watched = _load_watched()
        # Mise à jour si déjà présent
        for w in watched:
            if w["type"] == etype and w["value"] == value:
                if "notes" in body:
                    w["notes"] = str(body["notes"])[:500]
                _save_watched(watched)
                return jsonify({"ok": True, "action": "updated"})
        # Ajout
        entry = {
            "type": etype,
            "value": value,
            "added_at": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
            "notes": str(body.get("notes", ""))[:500],
        }
        watched.append(entry)
        _save_watched(watched)

    return jsonify({"ok": True, "action": "added"})


@app.route("/api/watched-entities", methods=["DELETE"])
def api_watched_delete():
    """Retire une entité de la surveillance (paramètres ?type=...&value=...)."""
    etype = (request.args.get("type") or "").strip().upper()
    value = (request.args.get("value") or "").strip()
    if not etype or not value:
        return jsonify({"error": "Paramètres type et value requis"}), 400

    with _watched_lock:
        watched = _load_watched()
        before = len(watched)
        watched = [w for w in watched if not (w["type"] == etype and w["value"] == value)]
        _save_watched(watched)

    return jsonify({"ok": True, "removed": len(watched) < before})


# ── Comparaison temporelle ────────────────────────────────────────────────────

@app.route("/api/analytics/compare")
def api_analytics_compare():
    """Compare deux périodes temporelles.

    Paramètres :
      from1, to1 : première période (YYYY-MM-DD)
      from2, to2 : deuxième période (YYYY-MM-DD)
    """
    from collections import defaultdict, Counter

    from1 = request.args.get("from1", "").strip()
    to1   = request.args.get("to1",   "").strip()
    from2 = request.args.get("from2", "").strip()
    to2   = request.args.get("to2",   "").strip()

    if not (from1 and to1 and from2 and to2):
        return jsonify({"error": "Paramètres from1, to1, from2, to2 requis"}), 400

    def _in_range(date_str: str, d_from: str, d_to: str) -> bool:
        d = date_str[:10] if date_str else ""
        return bool(d and d_from <= d <= d_to)

    def _stats(articles):
        if not articles:
            return {"count": 0, "sentiment": {}, "top_sources": [], "top_entities": []}
        sentiments = Counter(a.get("sentiment", "") for a in articles if a.get("sentiment"))
        sources = Counter(a.get("Sources", "") for a in articles if a.get("Sources"))
        entities: dict = defaultdict(Counter)
        for a in articles:
            ents = a.get("entities")
            if not isinstance(ents, dict):
                continue
            for etype, vals in ents.items():
                if isinstance(vals, list):
                    for v in vals:
                        if isinstance(v, str) and v.strip():
                            entities[etype][v.strip()] += 1
        top_entities = []
        for etype, counts in entities.items():
            for val, cnt in counts.most_common(5):
                top_entities.append({"type": etype, "value": val, "count": cnt})
        top_entities.sort(key=lambda x: x["count"], reverse=True)
        return {
            "count": len(articles),
            "sentiment": dict(sentiments),
            "top_sources": [{"source": s, "count": c} for s, c in sources.most_common(5)],
            "top_entities": top_entities[:20],
        }

    all_articles = []
    for data_dir in [PROJECT_ROOT / "data" / "articles", PROJECT_ROOT / "data" / "articles-from-rss"]:
        if not data_dir.exists():
            continue
        for json_file in data_dir.rglob("*.json"):
            if "cache" in str(json_file):
                continue
            try:
                arts = json.loads(json_file.read_text(encoding="utf-8", errors="replace"))
                if isinstance(arts, list):
                    all_articles.extend(arts)
            except (json.JSONDecodeError, OSError):
                continue

    # Déduplication par URL
    seen = set()
    deduped = []
    for a in all_articles:
        url = a.get("URL", "")
        if url and url in seen:
            continue
        if url:
            seen.add(url)
        deduped.append(a)

    p1 = [a for a in deduped if _in_range(a.get("Date de publication", ""), from1, to1)]
    p2 = [a for a in deduped if _in_range(a.get("Date de publication", ""), from2, to2)]

    return jsonify({
        "period1": {"from": from1, "to": to1, **_stats(p1)},
        "period2": {"from": from2, "to": to2, **_stats(p2)},
    })


# ── Export CSV / XLSX ─────────────────────────────────────────────────────────

@app.route("/api/export/csv")
def api_export_csv():
    """Exporte un fichier JSON d'articles en CSV.

    Paramètre :
      path : chemin relatif du fichier JSON (ex: data/articles-from-rss/OpenAI.json)
    """
    import csv
    import io

    path = request.args.get("path", "").strip()
    if not path:
        abort(400, "Paramètre path requis")
    f = safe_path(path)

    try:
        articles = json.loads(f.read_text(encoding="utf-8", errors="replace"))
        if not isinstance(articles, list):
            abort(400, "Le fichier ne contient pas une liste d'articles")
    except json.JSONDecodeError as e:
        abort(400, f"JSON invalide : {e}")

    # Colonnes exportées
    FIELDS = ["Date de publication", "Sources", "URL", "Résumé",
              "sentiment", "score_sentiment", "ton_editorial", "score_ton", "score_pertinence"]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=FIELDS, extrasaction="ignore",
                             lineterminator="\n")
    writer.writeheader()
    for art in articles:
        row = {k: art.get(k, "") for k in FIELDS}
        # Aplatir les entités en chaîne
        entities = art.get("entities", {})
        if isinstance(entities, dict):
            row["entities"] = "; ".join(
                f"{etype}:{','.join(str(v) for v in vals)}"
                for etype, vals in entities.items() if isinstance(vals, list)
            )
        writer.writerow(row)

    csv_content = output.getvalue()
    stem = Path(path).stem
    return Response(
        csv_content,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{stem}.csv"'},
    )


@app.route("/api/export/xlsx")
def api_export_xlsx():
    """Exporte un fichier JSON d'articles en XLSX (Excel).

    Paramètre :
      path : chemin relatif du fichier JSON
    """
    path = request.args.get("path", "").strip()
    if not path:
        abort(400, "Paramètre path requis")
    f = safe_path(path)

    try:
        articles = json.loads(f.read_text(encoding="utf-8", errors="replace"))
        if not isinstance(articles, list):
            abort(400, "Le fichier ne contient pas une liste d'articles")
    except json.JSONDecodeError as e:
        abort(400, f"JSON invalide : {e}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Articles"

        FIELDS = ["Date de publication", "Sources", "URL", "Résumé",
                  "sentiment", "score_sentiment", "ton_editorial", "score_ton",
                  "score_pertinence", "Entités"]

        # En-tête
        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(bold=True, color="FFFFFF")
        for col, field in enumerate(FIELDS, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.fill = header_fill
            cell.font = header_font

        # Données
        for row_idx, art in enumerate(articles, 2):
            entities = art.get("entities", {})
            entity_str = ""
            if isinstance(entities, dict):
                entity_str = "; ".join(
                    f"{et}:{','.join(str(v) for v in vals)}"
                    for et, vals in entities.items() if isinstance(vals, list)
                )
            values = [
                art.get("Date de publication", ""),
                art.get("Sources", ""),
                art.get("URL", ""),
                art.get("Résumé", ""),
                art.get("sentiment", ""),
                art.get("score_sentiment", ""),
                art.get("ton_editorial", ""),
                art.get("score_ton", ""),
                art.get("score_pertinence", ""),
                entity_str,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if col == 4:  # Résumé
                    cell.alignment = Alignment(wrap_text=True)

        # Largeurs de colonnes
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 80
        ws.row_dimensions[1].height = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        stem = Path(path).stem
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'},
        )
    except ImportError:
        # Fallback : retourne un CSV si openpyxl n'est pas installé
        return api_export_csv()
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Variables d'environnement ─────────────────────────────────────────────────
# Lit et écrit le fichier .env à la racine du projet.
# Les valeurs des variables sensibles (bearer, PASSWORD…) sont masquées en lecture.

_ENV_FILE = PROJECT_ROOT / ".env"
_SENSITIVE_KEYS = {"bearer", "SMTP_PASSWORD", "NTFY_TOKEN", "ANTHROPIC_API_KEY"}
_READONLY_KEYS = set()  # clés qu'on refuse de modifier


def _parse_env_file(path: Path) -> list[dict]:
    """Parse .env ligne par ligne → [{key, value, masked, comment, raw}]."""
    entries = []
    if not path.exists():
        return entries
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            entries.append({"key": None, "value": None, "raw": line, "comment": True})
            continue
        if "=" in stripped:
            key, _, val = stripped.partition("=")
            key = key.strip()
            val = val.strip()
            # Supprimer les guillemets éventuels
            if (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
                val = val[1:-1]
            masked = key.lower() in {k.lower() for k in _SENSITIVE_KEYS}
            entries.append({"key": key, "value": val, "masked": masked, "comment": False})
        else:
            entries.append({"key": None, "value": None, "raw": line, "comment": True})
    return entries


def _serialize_env(entries: list[dict]) -> str:
    """Reconstruit le contenu .env depuis la liste d'entrées."""
    lines = []
    for e in entries:
        if e.get("comment"):
            lines.append(e.get("raw", ""))
        else:
            key = e["key"]
            val = e.get("value", "")
            # Mettre entre guillemets si la valeur contient des espaces ou des caractères spéciaux
            if " " in val or "#" in val or ";" in val:
                val = f'"{val}"'
            lines.append(f"{key}={val}")
    return "\n".join(lines) + "\n"


@app.route("/api/env", methods=["GET"])
def api_env_get():
    """Retourne la liste des variables d'environnement depuis .env.

    Les valeurs des clés sensibles sont masquées (remplacées par '***').
    """
    entries = _parse_env_file(_ENV_FILE)
    result = []
    for e in entries:
        if e.get("comment"):
            result.append({"type": "comment", "raw": e.get("raw", "")})
        else:
            display_val = "***" if e.get("masked") else e["value"]
            result.append({
                "type": "var",
                "key": e["key"],
                "value": display_val,
                "masked": e.get("masked", False),
            })
    return jsonify(result)


@app.route("/api/env", methods=["POST"])
def api_env_post():
    """Crée ou met à jour une variable dans .env.

    Body JSON : { key: str, value: str }
    """
    body = request.get_json(force=True, silent=True) or {}
    key = (body.get("key") or "").strip()
    value = str(body.get("value") or "")

    if not key or not key.replace("_", "").isalnum():
        return jsonify({"error": "Clé invalide (alphanumérique + underscore uniquement)"}), 400

    entries = _parse_env_file(_ENV_FILE)

    # Mise à jour si la clé existe déjà
    found = False
    for e in entries:
        if not e.get("comment") and e.get("key") == key:
            e["value"] = value
            found = True
            break

    if not found:
        entries.append({"key": key, "value": value, "masked": False, "comment": False})

    # Sauvegarde atomique
    tmp = _ENV_FILE.with_suffix(".env.tmp")
    tmp.write_text(_serialize_env(entries), encoding="utf-8")
    tmp.replace(_ENV_FILE)

    # Recharger dans l'environnement courant du processus Flask
    os.environ[key] = value

    # Invalider le singleton Config pour que les prochains appels get_config()
    # voient les nouvelles valeurs (ANTHROPIC_API_KEY, CLAUDE_MODEL_*, etc.)
    try:
        from utils.config import get_config as _get_config
        _get_config(force_reload=True)
    except Exception:
        pass

    return jsonify({"ok": True, "key": key})


@app.route("/api/env/<key>", methods=["DELETE"])
def api_env_delete(key: str):
    """Supprime une variable de .env."""
    if not key or not key.replace("_", "").isalnum():
        return jsonify({"error": "Clé invalide"}), 400

    entries = _parse_env_file(_ENV_FILE)
    entries = [e for e in entries if e.get("comment") or e.get("key") != key]

    tmp = _ENV_FILE.with_suffix(".env.tmp")
    tmp.write_text(_serialize_env(entries), encoding="utf-8")
    tmp.replace(_ENV_FILE)

    os.environ.pop(key, None)
    try:
        from utils.config import get_config as _get_config
        _get_config(force_reload=True)
    except Exception:
        pass
    return jsonify({"ok": True, "key": key})


# ── Test de connexion IA ───────────────────────────────────────────────────────

@app.route("/api/ai-check", methods=["POST"])
def api_ai_check():
    """Vérifie la connexion à un fournisseur IA en envoyant un prompt minimal.

    Body JSON : { "provider": "euria" | "claude" }
    Retourne  : { ok: bool, message: str, latency_ms: int }
    """
    import time as _time
    body = request.get_json(force=True, silent=True) or {}
    provider = (body.get("provider") or "").strip().lower()
    if provider not in ("euria", "claude"):
        return jsonify({"error": "provider doit être 'euria' ou 'claude'"}), 400

    try:
        from utils.api_client import EurIAClient, ClaudeClient
    except ImportError as e:
        return jsonify({"ok": False, "message": f"Import impossible : {e}", "latency_ms": 0}), 500

    prompt = "Réponds uniquement par 'OK'."
    t0 = _time.monotonic()
    try:
        if provider == "euria":
            url = os.environ.get("URL", "").strip()
            bearer = os.environ.get("bearer", "").strip()
            if not url or not bearer:
                return jsonify({"ok": False, "message": "URL ou bearer non configuré.", "latency_ms": 0})
            client = EurIAClient(url=url, bearer=bearer)
            result = client.ask(prompt, timeout=10)
        else:
            api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
            if not api_key:
                return jsonify({"ok": False, "message": "ANTHROPIC_API_KEY non configurée.", "latency_ms": 0})
            client = ClaudeClient(api_key=api_key)
            result = client.ask(prompt, max_tokens=16, timeout=10)

        latency_ms = int((_time.monotonic() - t0) * 1000)
        ok = bool(result and len(result.strip()) > 0)
        return jsonify({"ok": ok, "message": result.strip() if ok else "Réponse vide.", "latency_ms": latency_ms})

    except Exception as exc:
        latency_ms = int((_time.monotonic() - t0) * 1000)
        return jsonify({"ok": False, "message": str(exc), "latency_ms": latency_ms})


# ── Rafraîchissement d'un résumé article ─────────────────────────────────────

@app.route("/api/article/refresh-resume", methods=["POST"])
def api_article_refresh_resume():
    """Régénère le résumé d'un article via l'IA choisie et met à jour le fichier JSON.

    Body JSON :
      file_path   (str) — chemin relatif du fichier JSON dans PROJECT_ROOT
      article_url (str) — URL de l'article à rafraîchir
      provider    (str) — 'euria', 'claude', ou 'auto' (utilise AI_PROVIDER depuis .env)
    Retourne : { ok: bool, resume: str }
    """
    body = request.get_json(force=True, silent=True) or {}
    rel_path = (body.get("file_path") or "").strip()
    article_url = (body.get("article_url") or "").strip()
    provider = (body.get("provider") or "auto").strip().lower()

    if not rel_path or not article_url:
        return jsonify({"error": "file_path et article_url sont requis"}), 400

    # Validation du chemin — uniquement data/ et articles-from-rss/
    if not (rel_path.startswith("data/") or rel_path.startswith("samples/")):
        return jsonify({"error": "Chemin non autorisé"}), 403

    target = (PROJECT_ROOT / rel_path).resolve()
    if not str(target).startswith(str(PROJECT_ROOT) + "/"):
        return jsonify({"error": "Accès refusé"}), 403
    if not target.exists():
        return jsonify({"error": "Fichier non trouvé"}), 404

    try:
        articles = json.loads(target.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        return jsonify({"error": f"Lecture impossible : {e}"}), 500

    # Trouver l'article par URL
    article = next((a for a in articles if a.get("URL") == article_url), None)
    if article is None:
        return jsonify({"error": "Article non trouvé dans le fichier"}), 404

    # Récupérer le texte source : tenter de re-fetcher l'article original depuis son URL,
    # fallback sur le résumé existant (re-résumer un résumé dégrade la qualité mais reste utile).
    source_text = ""
    original_url = article.get("URL", "").strip()
    if original_url:
        try:
            import requests as _req
            from bs4 import BeautifulSoup as _BS
            resp = _req.get(original_url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
            soup = _BS(resp.text, "html.parser")
            for tag in soup(["script", "style", "nav", "footer", "header"]):
                tag.decompose()
            source_text = " ".join(soup.get_text(separator=" ").split())[:15000]
        except Exception:
            pass  # fallback ci-dessous
    if not source_text.strip():
        source_text = article.get("Résumé") or article.get("Titre") or ""
    if not source_text.strip():
        return jsonify({"error": "Aucun texte source disponible pour générer un résumé"}), 400

    # Sélectionner le client IA
    try:
        from utils.api_client import EurIAClient, ClaudeClient
        if provider == "auto":
            provider = os.environ.get("AI_PROVIDER", "euria").strip().lower()

        if provider == "claude":
            api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
            if not api_key:
                return jsonify({"error": "ANTHROPIC_API_KEY non configurée"}), 400
            client = ClaudeClient(api_key=api_key)
        else:
            url_env = os.environ.get("URL", "").strip()
            bearer = os.environ.get("bearer", "").strip()
            if not url_env or not bearer:
                return jsonify({"error": "URL ou bearer non configuré"}), 400
            client = EurIAClient(url=url_env, bearer=bearer)

        new_resume = client.generate_summary(source_text)
    except Exception as exc:
        return jsonify({"error": f"Erreur IA : {exc}"}), 500

    # Mettre à jour le fichier JSON
    article["Résumé"] = new_resume
    try:
        tmp = target.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(articles, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(target)
    except OSError as e:
        return jsonify({"error": f"Erreur écriture : {e}"}), 500

    return jsonify({"ok": True, "resume": new_resume})


# ── Vérification répertoire backup ────────────────────────────────────────────

@app.route("/api/backup/check-dir", methods=["POST"])
def api_backup_check_dir():
    """Vérifie qu'un répertoire existe et est accessible en écriture.

    Body JSON : { "path": str }
    Retourne  : { ok: bool, message: str }
    """
    body = request.get_json(force=True, silent=True) or {}
    path_str = (body.get("path") or "").strip()
    if not path_str:
        return jsonify({"ok": False, "message": "Chemin vide"}), 400

    import stat as _stat
    p = Path(path_str)
    try:
        if p.exists():
            if not p.is_dir():
                return jsonify({"ok": False, "message": "Ce chemin n'est pas un répertoire"})
            if not os.access(str(p), os.W_OK):
                return jsonify({"ok": False, "message": "Répertoire non inscriptible"})
            try:
                usage = _stat.os.statvfs(str(p))
                free_gb = (usage.f_bavail * usage.f_frsize) / (1024 ** 3)
                return jsonify({"ok": True, "message": f"Accessible · {free_gb:.1f} Go libres"})
            except Exception:
                return jsonify({"ok": True, "message": "Accessible en écriture"})
        else:
            # Le répertoire n'existe pas encore — vérifier que le parent est accessible
            parent = p.parent
            if not parent.exists():
                return jsonify({"ok": False, "message": f"Répertoire parent introuvable : {parent}"})
            if not os.access(str(parent), os.W_OK):
                return jsonify({"ok": False, "message": f"Répertoire parent non inscriptible : {parent}"})
            return jsonify({"ok": True, "message": "Répertoire sera créé automatiquement"})
    except Exception as exc:
        return jsonify({"ok": False, "message": str(exc)})


# ── Disponibilité des fournisseurs IA (pour le frontend) ──────────────────────

@app.route("/api/ai-providers")
def api_ai_providers():
    """Retourne la liste des fournisseurs IA dont les credentials sont configurés.

    Retourne : { providers: ["euria"|"claude", ...], active: str }
    """
    available = []
    if os.environ.get("URL", "").strip() and os.environ.get("bearer", "").strip():
        available.append("euria")
    if os.environ.get("ANTHROPIC_API_KEY", "").strip():
        available.append("claude")
    active = os.environ.get("AI_PROVIDER", "euria").strip().lower()
    return jsonify({"providers": available, "active": active})


# ── Clustering thématique ──────────────────────────────────────────────────────

@app.route("/api/analytics/clusters")
def api_analytics_clusters():
    """Retourne les clusters thématiques des N derniers jours.

    Query params:
      days      : fenêtre temporelle en jours (défaut 7)
      min_size  : taille minimale d'un cluster (défaut 2)
    """
    days = max(1, min(int(request.args.get("days", 7)), 365))
    min_size = max(1, int(request.args.get("min_size", 2)))

    try:
        _sys.path.insert(0, str(PROJECT_ROOT))
        from scripts.cluster_articles import load_articles, cluster_articles

        articles = load_articles(PROJECT_ROOT, days=days)
        clusters = cluster_articles(articles)
        clusters = [c for c in clusters if c["count"] >= min_size]

        return jsonify({
            "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "window_days": days,
            "total_articles": len(articles),
            "clusters": clusters,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ── Chatbot IA ────────────────────────────────────────────────────────────────

# Paramètres du chatbot
_CHAT_MAX_CONTEXT_FILES  = 10    # Nombre maximum de fichiers de contexte par requête
_CHAT_MAX_CONTEXT_CHARS  = 12000 # Taille maximale (caractères) par fichier de contexte

def _build_notes_context(period: str = "week") -> str:
    """Génère un bloc de contexte Markdown à partir des annotations personnelles.

    Args:
        period: "week" (7 derniers jours), "month" (30 jours), "all" (tout)

    Returns:
        Texte Markdown formaté listant les notes personnelles avec métadonnées article.
    """
    annotations = _load_annotations()
    if not annotations:
        return ""

    # Déterminer la date de début selon la période
    now = datetime.datetime.now(datetime.timezone.utc)
    if period == "week":
        cutoff = now - datetime.timedelta(days=7)
        period_label = "7 derniers jours"
    elif period == "month":
        cutoff = now - datetime.timedelta(days=30)
        period_label = "30 derniers jours"
    else:
        cutoff = None
        period_label = "toutes les notes"

    # Filtrer les annotations ayant une note ou des tags, et selon la période
    selected = {}
    for url, ann in annotations.items():
        has_content = ann.get("notes", "").strip() or [t for t in (ann.get("tags") or []) if t]
        if not has_content:
            continue
        if cutoff is not None:
            updated_raw = ann.get("updated_at", "")
            if updated_raw:
                try:
                    # Supporte ISO 8601 avec ou sans timezone
                    updated = datetime.datetime.fromisoformat(updated_raw.replace("Z", "+00:00"))
                    if updated.tzinfo is None:
                        updated = updated.replace(tzinfo=datetime.timezone.utc)
                    if updated < cutoff:
                        continue
                except Exception:
                    pass  # Conserver si la date est illisible
        selected[url] = ann

    if not selected:
        return f"*Aucune note personnelle pour la période : {period_label}.*"

    # Construire un index article {url: article_dict} pour enrichir avec le titre/source
    article_index: dict = {}

    def _index_articles(directory: Path) -> None:
        if not directory.exists():
            return
        for f in directory.rglob("*.json"):
            parts = f.relative_to(PROJECT_ROOT).parts
            if "cache" in parts:
                continue
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    for a in data:
                        if isinstance(a, dict) and a.get("URL"):
                            article_index[a["URL"]] = a
            except Exception:
                pass

    _index_articles(PROJECT_ROOT / "data" / "articles-from-rss")
    _index_articles(PROJECT_ROOT / "data" / "articles")

    # Trier par date de mise à jour décroissante
    sorted_entries = sorted(
        selected.items(),
        key=lambda kv: kv[1].get("updated_at", ""),
        reverse=True,
    )

    lines = [
        f"## Notes personnelles de lecture ({period_label})",
        f"*{len(sorted_entries)} note(s) trouvée(s)*",
        "",
    ]

    for url, ann in sorted_entries:
        notes = ann.get("notes", "").strip()
        tags = [t for t in (ann.get("tags") or []) if t]
        is_important = bool(ann.get("is_important", False))
        updated_at = (ann.get("updated_at") or "")[:10]
        article = article_index.get(url, {})
        source = article.get("Sources", "")
        pub_date = (article.get("Date de publication") or "")[:10]

        # Extraire le titre de l'article
        titre = (article.get("Titre") or "").strip()
        if not titre:
            for line in (article.get("Résumé") or "").split("\n"):
                line = line.strip().lstrip("*_#").strip()
                if len(line) > 10:
                    titre = line[:120]
                    break
        if not titre:
            titre = url.rstrip("/").split("/")[-1][:80] or "Sans titre"

        star = "⭐ " if is_important else ""
        meta_parts = []
        if pub_date:
            meta_parts.append(pub_date)
        if source:
            meta_parts.append(source)
        meta = " · ".join(meta_parts)

        lines.append(f"### {star}{titre}")
        if meta:
            lines.append(f"*{meta}*")
        lines.append(f"URL : {url}")
        if updated_at:
            lines.append(f"Note ajoutée le : {updated_at}")
        if tags:
            lines.append(f"Tags : {', '.join(tags)}")
        if notes:
            lines.append(f"\n> {notes}")
        lines.append("")

    return "\n".join(lines)


@app.route("/api/chat/stream", methods=["POST"])
def api_chat_stream():
    """Chatbot IA en streaming SSE.

    Body JSON :
      messages      (list)   — historique de conversation [{ role, content }, ...]
      context_files (list)   — chemins relatifs des fichiers à inclure comme contexte (optionnel)
      notes_period  (string) — période des notes personnelles à inclure : "week", "month" ou "all" (optionnel)

    Retourne un flux SSE au format OpenAI : data: {"choices":[{"delta":{"content":"..."},...}]}
    """
    import requests as req

    body = request.get_json(force=True, silent=True) or {}
    messages        = body.get("messages", [])
    context_files   = body.get("context_files", [])
    notes_period    = body.get("notes_period", None)
    # Contexte entité pré-formaté (texte brut) fourni par Terminal IA depuis EntityArticlePanel
    entity_context  = body.get("entity_context", "").strip()
    # Permet au frontend de choisir le provider pour cette requête.
    # Valeurs acceptées : "euria" | "claude". Sinon, fallback sur AI_PROVIDER env.
    provider_override = body.get("provider", "").strip().lower()

    if not messages:
        return jsonify({"error": "messages est requis"}), 400

    # Valider et charger les fichiers de contexte
    context_blocks = []
    for rel in context_files[:_CHAT_MAX_CONTEXT_FILES]:  # Limite au nombre maximal configuré
        rel = str(rel).strip()
        if not rel:
            continue
        # Restriction aux répertoires autorisés
        if not (rel.startswith("data/") or rel.startswith("rapports/") or rel.startswith("samples/")):
            continue
        target = (PROJECT_ROOT / rel).resolve()
        if not str(target).startswith(str(PROJECT_ROOT) + "/"):
            continue
        if not target.exists() or not target.is_file():
            continue
        try:
            raw = target.read_text(encoding="utf-8")
            # Tronquer les fichiers volumineux
            if len(raw) > _CHAT_MAX_CONTEXT_CHARS:
                raw = raw[:_CHAT_MAX_CONTEXT_CHARS] + "\n…[tronqué]"
            ext = target.suffix.lower()
            lang = "json" if ext == ".json" else "markdown"
            context_blocks.append(f"### Fichier : {rel}\n```{lang}\n{raw}\n```")
        except OSError:
            continue

    # Charger les notes personnelles si demandé
    notes_block = None
    if notes_period and notes_period in ("week", "month", "all"):
        with _annotations_lock:
            notes_text = _build_notes_context(notes_period)
        if notes_text:
            notes_block = notes_text

    # Construire le message système
    from datetime import datetime as _dt
    _today = _dt.now().strftime("%A %d %B %Y")
    system_parts = [
        "Tu es un assistant IA intégré à WUDD.ai, une plateforme de veille de presse en français.",
        f"La date d'aujourd'hui est le {_today}. Tiens-en compte pour contextualiser toutes tes réponses sur l'actualité.",
        "Tu aides l'utilisateur à analyser des articles de presse, des rapports et des données JSON.",
        "Tu peux produire des tableaux Markdown, des résumés, des analyses comparatives.",
        "Réponds toujours en français, de manière concise et structurée.",
        "Utilise du Markdown pour les tableaux, listes et mise en forme.",
        "Ne génère pas de balises <think>.",
        "IMPORTANT — Tu es un assistant en LECTURE SEULE. Tu ne peux PAS supprimer, effacer, modifier ou détruire des fichiers, des données ou des rapports.",
        "Si l'utilisateur te demande de supprimer ou d'effacer des fichiers, des données ou des rapports"
        " (quelle que soit la formulation : commandes shell, appels API, code, instructions, etc.),"
        " refuse poliment et rappelle-lui que cette opération est impossible depuis ce chatbot.",
    ]
    if entity_context:
        system_parts.append("\n\n## Contexte entité (données WUDD.ai) :\n")
        system_parts.append(entity_context)
    if notes_block:
        system_parts.append("\n\n## Notes personnelles de l'utilisateur :\n")
        system_parts.append(notes_block)
    if context_blocks:
        system_parts.append("\n\n## Fichiers de contexte fournis par l'utilisateur :\n")
        system_parts.extend(context_blocks)

    system_prompt = "\n".join(system_parts)

    # Nettoyer les messages : ne garder que role + content (strip les champs frontend
    # comme "streaming" ou "error" qui feraient échouer l'API Anthropic avec 400).
    # Ignorer aussi les messages avec contenu vide (erreurs mid-stream).
    clean_messages = [
        {"role": m["role"], "content": m["content"]}
        for m in messages
        if m.get("role") in ("user", "assistant") and (m.get("content") or "").strip()
    ]

    # Construire les messages complets (système + historique)
    full_messages = [{"role": "system", "content": system_prompt}] + clean_messages

    provider = provider_override if provider_override in ("euria", "claude") \
               else os.environ.get("AI_PROVIDER", "euria").strip().lower()

    if provider == "claude":
        api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
        if not api_key:
            return jsonify({"error": "ANTHROPIC_API_KEY manquante dans .env (AI_PROVIDER=claude)"}), 503
        # Routage Haiku/Sonnet selon la taille du contexte :
        # - contexte court (< 3 000 chars) → Haiku (5-8× moins cher, suffisant pour Q&R simples)
        # - contexte long ou analyse multi-articles → Sonnet
        total_context_chars = sum(len(m.get("content", "")) for m in clean_messages)
        if total_context_chars < 3000:
            model = os.environ.get("CLAUDE_MODEL_BATCH", "claude-haiku-4-5-20251001")
        else:
            model = os.environ.get("CLAUDE_MODEL_SYNTHESIS", "claude-sonnet-4-6")
        # Claude ne supporte pas role=system dans messages[], on extrait le system
        claude_messages = [m for m in full_messages if m["role"] != "system"]
        from utils.api_client import ClaudeClient as _ClaudeClient
        _claude = _ClaudeClient(api_key=api_key)

        def generate_chat():
            yield from _claude.stream(
                prompt="",
                model=model,
                system=system_prompt,
                max_tokens=4096,
                timeout=180,
                messages=claude_messages,
            )

    else:
        api_url = os.environ.get("URL", "").strip()
        bearer  = os.environ.get("bearer", "").strip()
        if not api_url or not bearer:
            return jsonify({"error": "URL ou bearer manquant dans .env (AI_PROVIDER=euria)"}), 503
        payload = {
            "messages": full_messages,
            "model": "qwen3",
            "stream": True,
            "enable_web_search": True,
        }
        api_headers = {
            "Authorization": f"Bearer {bearer}",
            "Content-Type": "application/json",
        }

        def generate_chat():
            try:
                r = req.post(api_url, json=payload, headers=api_headers, stream=True, timeout=180)
                r.raise_for_status()
                for line in r.iter_lines():
                    if line:
                        decoded = line.decode("utf-8")
                        if not decoded.startswith("data:"):
                            decoded = "data: " + decoded
                        yield decoded + "\n\n"
            except Exception as exc:
                yield f'data: {json.dumps({"error": str(exc)})}\n\n'

    return Response(
        stream_with_context(generate_chat()),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/chat/save", methods=["POST"])
def api_chat_save():
    """Sauvegarde une conversation ou une réponse IA en Markdown dans rapports/.

    Body JSON :
      content  (str)  — contenu Markdown à sauvegarder
      filename (str)  — nom de fichier suggéré (sans extension, optionnel)
      subdir   (str)  — sous-répertoire dans rapports/ (défaut : "_WUDD.AI_")

    Retourne : { ok: bool, path: str }
    """
    body = request.get_json(force=True, silent=True) or {}
    content  = (body.get("content") or "").strip()
    filename = (body.get("filename") or "").strip()
    subdir   = (body.get("subdir") or "_WUDD.AI_").strip()

    if not content:
        return jsonify({"error": "content est requis"}), 400

    # Sanitiser le sous-répertoire
    subdir = re.sub(r"[^\w\-/]", "_", subdir).strip("/")
    if not subdir:
        subdir = "_WUDD.AI_"

    # Sanitiser le nom de fichier
    if filename:
        filename = re.sub(r"[^\w\-]", "_", filename)[:80]
    else:
        filename = "chat"

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = f"{filename}_{ts}.md"

    save_dir = (PROJECT_ROOT / "rapports" / subdir).resolve()
    # Vérifier que le répertoire cible reste dans rapports/
    rapports_root = (PROJECT_ROOT / "rapports").resolve()
    if not str(save_dir).startswith(str(rapports_root)):
        return jsonify({"error": "Répertoire non autorisé"}), 403

    try:
        save_dir.mkdir(parents=True, exist_ok=True)
        out_path = save_dir / safe_name
        out_path.write_text(content, encoding="utf-8")
        rel = str(out_path.relative_to(PROJECT_ROOT)).replace("\\", "/")
        return jsonify({"ok": True, "path": rel})
    except OSError as e:
        return jsonify({"error": f"Erreur écriture : {e}"}), 500


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_app(path):
    dist = Path(__file__).parent / "dist"
    if not dist.exists():
        return (
            "<h1>Frontend non compilé</h1>"
            "<p>Exécutez <code>npm run build</code> dans le dossier <code>viewer/</code></p>",
            503,
        )
    target = dist / path
    if path and target.exists() and target.is_file():
        return send_from_directory(str(dist), path)
    # SPA fallback : toutes les routes renvoient index.html (sans cache)
    response = send_from_directory(str(dist), "index.html")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


if __name__ == "__main__":
    print(f"WUDD.ai Viewer — racine projet : {PROJECT_ROOT}")
    print("API disponible sur http://localhost:5050/api/files")
    app.run(host="0.0.0.0", port=5050, debug=False, threaded=True)
